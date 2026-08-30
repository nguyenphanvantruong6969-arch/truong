"""Sinh bộ DỮ LIỆU MÔ PHỎNG để chạy thử phần mềm ở quy mô thật.

    ./.venv/bin/python du_lieu_test/tao_du_lieu_test.py

⚠️  ĐÂY LÀ DỮ LIỆU BỊA, KHÔNG PHẢI HỌC SINH CÓ THẬT.
    Tên và mã đều do máy sinh ngẫu nhiên (có seed nên lần nào cũng ra
    đúng bộ đó). Trình bày như số liệu khảo sát thật là bịa đặt dữ liệu.

Sinh hai bộ:
  BỘ SẠCH   — 120 học sinh, 10 CLB, không lỗi. Dùng để chạy thử trọn
              quy trình từ nạp file tới xuất kết quả.
  BỘ CÓ LỖI — cố ý cài sẵn 5 tình huống sai, để kiểm tra phần mềm có
              thật sự cảnh báo hay không.
"""

import os
import random

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

THU_MUC = os.path.dirname(os.path.abspath(__file__))
SEED = 2026
XANH = PatternFill("solid", fgColor="DCE9E0")
VANG = PatternFill("solid", fgColor="F7EEDA")

# --------------------------------------------------------------- #
# 10 CLB — tổng 130 suất cho 120 em.
#
# Bộ này được CỐ Ý THIẾT KẾ CHO CẠNH TRANH CAO, không phải mô phỏng một
# phân bố nguyện vọng tự nhiên. Lý do: với nguyện vọng rải đều thì gần
# như em nào cũng được nguyện vọng 1 và suất dự trữ không bao giờ phải
# dùng tới — nhìn bảng kết quả không thấy thuật toán làm gì cả.
#
# Ba điều làm cơ chế lộ ra:
#   1. Nhu cầu dồn vào 3 CLB "hot" (trọng số 10) và né 3 CLB "nguội"
#      (trọng số 0.6) -> có CLB chật cứng bên cạnh CLB thừa chỗ.
#   2. Suất dự trữ đặt ĐÚNG ở các CLB chật. Đặt ở CLB còn chỗ thì suất
#      dự trữ vô dụng, vì ai cũng vào được.
#   3. Điểm của nhóm dự trữ lệch thấp hơn -> nhiều em vào được DÙ điểm
#      dưới mức chuẩn chung, và nhìn bảng là thấy ngay.
# --------------------------------------------------------------- #
CLB = [
    # (mã, tên, chỉ tiêu, chỉ tiêu dự trữ, nhóm dự trữ, độ hút)
    ("clb_bongda",    "CLB Bóng đá",           20, 4, "chinh_sach", 10.0),
    ("clb_tinhoc",    "CLB Tin học",           12, 3, "chinh_sach", 10.0),
    ("clb_mythuat",   "CLB Mỹ thuật",          12, 3, "khoi_10",    10.0),
    ("clb_bongro",    "CLB Bóng rổ",           18, 0, "",            3.0),
    ("clb_amnhac",    "CLB Âm nhạc",           14, 0, "",            3.0),
    ("clb_tienganh",  "CLB Tiếng Anh",         16, 2, "khoi_10",    10.0),
    ("clb_robotics",  "CLB Robotics",           8, 0, "",            2.0),
    ("clb_khoahoc",   "CLB Khoa học",           8, 0, "",            0.75),
    ("clb_vanhoc",    "CLB Văn học",           10, 0, "",            0.75),
    ("clb_tinhnguyen","CLB Tình nguyện",       12, 0, "",            0.75),
]

# Điểm: nhóm dự trữ lệch thấp hơn nhóm thường. Đây KHÔNG phải nhận định
# về học sinh diện chính sách — đây là cách dựng số liệu sao cho suất dự
# trữ thực sự phải làm việc, để người chạy thử nhìn thấy cơ chế.
DIEM_TB = {"": 7.6, "chinh_sach": 6.3, "khoi_10": 6.5}
DIEM_LECH = 1.15


HO = ["Nguyễn", "Trần", "Lê", "Phạm", "Hoàng", "Huỳnh", "Phan", "Vũ",
      "Võ", "Đặng", "Bùi", "Đỗ", "Hồ", "Ngô", "Dương", "Lý"]
DEM_NAM = ["Văn", "Hữu", "Đức", "Minh", "Quang", "Thành", "Anh", "Bá"]
DEM_NU  = ["Thị", "Ngọc", "Thu", "Minh", "Phương", "Hoài", "Kim", "Diệu"]
TEN_NAM = ["An", "Bình", "Cường", "Dũng", "Đạt", "Giang", "Hải", "Hùng",
           "Khoa", "Long", "Minh", "Nam", "Phúc", "Quân", "Sơn", "Tuấn",
           "Việt", "Vinh", "Bảo", "Khang"]
TEN_NU  = ["Anh", "Chi", "Dung", "Hà", "Hạnh", "Hoa", "Hương", "Lan",
           "Linh", "Mai", "Nga", "Ngân", "Nhung", "Oanh", "Phương",
           "Quỳnh", "Thảo", "Trang", "Vân", "Yến"]


def sinh_ten(rng):
    nu = rng.random() < 0.5
    return "%s %s %s" % (
        rng.choice(HO),
        rng.choice(DEM_NU if nu else DEM_NAM),
        rng.choice(TEN_NU if nu else TEN_NAM),
    )


def ghi_sheet(ws, header, rows, to_mau=None):
    ws.append(header)
    for r in rows:
        ws.append(r)
    for i, ten_cot in enumerate(header, start=1):
        o = ws.cell(row=1, column=i)
        o.font = Font(bold=True)
        o.fill = to_mau or XANH
        o.alignment = Alignment(vertical="center")
        rong = max([len(str(ten_cot))] +
                   [len(str(r[i - 1])) for r in rows if i - 1 < len(r)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(rong + 3, 12), 34)
    ws.freeze_panes = "A2"


def them_sheet_ghi_chu(wb, dong):
    hd = wb.create_sheet("Ghi chú")
    hd.column_dimensions["A"].width = 82
    hd.append(["DỮ LIỆU MÔ PHỎNG — KHÔNG PHẢI HỌC SINH CÓ THẬT"])
    hd["A1"].font = Font(bold=True, size=13, color="A63A2B")
    hd.append([""])
    for d in dong:
        hd.append([d])
    hd.append([""])
    hd.append(["Sinh bằng du_lieu_test/tao_du_lieu_test.py, seed = %d" % SEED])
    hd.append(["Chạy lại script luôn cho ra đúng bộ dữ liệu này."])


def luu(wb, ten):
    p = os.path.join(THU_MUC, ten)
    wb.save(p)
    print("  da tao %-42s" % ten)
    return p


# =============================================================== #
# BỘ SẠCH
# =============================================================== #
def chon_theo_trong_so(rng, ung_vien, trong_so, n):
    """Bốc n phần tử KHÔNG lặp, xác suất tỉ lệ với trọng số."""
    con = list(ung_vien)
    ts = dict(trong_so)
    ra = []
    for _ in range(min(n, len(con))):
        tong = sum(ts[c] for c in con)
        moc = rng.uniform(0, tong)
        chay = 0.0
        for c in con:
            chay += ts[c]
            if chay >= moc:
                ra.append(c)
                con.remove(c)
                break
    return ra


def sinh_diem(rng, nhom):
    d = rng.gauss(DIEM_TB.get(nhom, DIEM_TB[""]), DIEM_LECH)
    return round(min(10.0, max(4.0, d)), 1)


def bo_sach():
    rng = random.Random(SEED)
    ma_clb = [c[0] for c in CLB]
    hut = {c[0]: c[5] for c in CLB}

    hoc_sinh = []
    for i in range(1, 121):
        ma = "HS%03d" % i
        ten = sinh_ten(rng)
        r = rng.random()
        nhom = "chinh_sach" if r < 0.13 else ("khoi_10" if r < 0.22 else "")
        hoc_sinh.append((ma, ten, nhom))

    # --- 1. Danh sách CLB ---
    wb = Workbook(); ws = wb.active; ws.title = "Danh sách CLB"
    ghi_sheet(ws,
              ["club_id", "name", "capacity", "reserve_capacity", "reserve_group"],
              [list(c[:5]) for c in CLB])
    them_sheet_ghi_chu(wb, [
        "10 câu lạc bộ, tổng 130 suất cho 120 học sinh.",
        "",
        "Bốn CLB có suất dự trữ, và ĐỀU là CLB đông người đăng ký:",
        "Bóng đá và Tin học dành cho nhóm chinh_sach; Mỹ thuật và Tiếng",
        "Anh dành cho nhóm khoi_10. Đặt suất dự trữ ở CLB còn thừa chỗ",
        "thì suất đó vô dụng, vì ai cũng vào được.",
        "",
        "NHẬP FILE NÀY TRƯỚC hai file học sinh — nếu không, mọi học sinh",
        "sẽ bị bỏ qua vì mã CLB chưa tồn tại.",
    ])
    luu(wb, "TEST_01_danh_sach_CLB.xlsx")

    # --- 2. Chọn CLB muốn thi + điểm chấm, và 3. Nguyện vọng ---
    dong_thi, dong_nv = [], []
    so_cot_thi = 5
    so_cot_nv = 5
    for ma, ten, nhom in hoc_sinh:
        # Nguyện vọng bốc theo ĐỘ HÚT -> dồn vào vài CLB.
        nv = chon_theo_trong_so(rng, ma_clb, hut, rng.randint(2, so_cot_nv))
        # Đăng ký thi ĐÚNG những CLB đã xếp nguyện vọng. Thi một CLB mà
        # không xếp nguyện vọng vào đó là lượt thi bỏ phí — phần mềm cảnh
        # báo đúng như vậy, nên bộ dữ liệu demo không nên tự tạo ra.
        thi = list(nv)[:so_cot_thi]

        o_thi = [ma, ten, nhom]
        for i in range(so_cot_thi):
            if i < len(thi):
                o_thi += [thi[i], sinh_diem(rng, nhom)]
            else:
                o_thi += ["", ""]
        dong_thi.append(o_thi)
        dong_nv.append([ma, ten, nhom] + nv + [""] * (so_cot_nv - len(nv)))

    cot_thi = ["student_id", "name", "reserve_group"]
    for i in range(1, so_cot_thi + 1):
        cot_thi += ["test_club_%d" % i, "score_%d" % i]

    wb = Workbook(); ws = wb.active; ws.title = "Chọn CLB muốn thi"
    ghi_sheet(ws, cot_thi, dong_thi)
    them_sheet_ghi_chu(wb, [
        "120 học sinh, mỗi em đăng ký thi 2-5 CLB.",
        "",
        "Cột score_1 đi kèm test_club_1, score_2 đi kèm test_club_2, và",
        "cứ thế — GHÉP THEO SỐ THỨ TỰ trong tên cột, không theo vị trí.",
        "Nạp file này là có luôn điểm chấm, không phải gõ tay 400 ô.",
        "",
        "Cột reserve_group: khoảng 22%% số em thuộc diện dự trữ",
        "(chinh_sach hoặc khoi_10), phần còn lại để trống.",
        "",
        "Ô trống ở các cột test_club_* và score_* là bình thường.",
    ])
    luu(wb, "TEST_02_chon_CLB_muon_thi.xlsx")

    wb = Workbook(); ws = wb.active; ws.title = "Xếp hạng nguyện vọng"
    ghi_sheet(ws,
              ["student_id", "name", "reserve_group"] +
              ["pref_%d" % i for i in range(1, so_cot_nv + 1)],
              dong_nv)
    them_sheet_ghi_chu(wb, [
        "Cùng 120 học sinh, xếp 2-5 nguyện vọng theo THỨ TỰ CỘT:",
        "pref_1 là nguyện vọng mong muốn nhất.",
        "",
        "Nhu cầu CỐ Ý dồn vào vài CLB: Bóng đá, Tin học và Mỹ thuật đông",
        "gấp nhiều lần chỉ tiêu, còn Khoa học, Văn học và Tình nguyện thì",
        "gần như không ai chọn. Nhờ vậy mới thấy thuật toán phải làm việc:",
        "nhiều em trượt xuống nguyện vọng 2-3, và suất dự trữ thực sự",
        "quyết định ai vào ai không.",
    ])
    luu(wb, "TEST_03_xep_hang_nguyen_vong.xlsx")
    return hoc_sinh


# =============================================================== #
# BỘ CÓ LỖI CỐ Ý
# =============================================================== #
def bo_co_loi():
    wb = Workbook(); ws = wb.active; ws.title = "Có lỗi cố ý"
    dong = [
        # (mã, tên, nhóm, nv1, nv2, điểm) — kèm chú thích ở sheet sau
        ["HS201", "Nguyễn Văn Một",   "",            "clb_bongro",  "clb_amnhac", ""],
        ["HS202", "Trần Thị Hai",     "",            "clb_bongro",  "",           ""],
        ["HS202", "Trần Thị Hai",     "",            "clb_amnhac",  "",           ""],
        ["hs201", "Nguyễn Văn Một",   "",            "clb_vanhoc",  "",           ""],
        ["HS204", "Lê Văn Bốn",       "chinh_sac",   "clb_tienganh","",           ""],
        ["HS205", "Phạm Thị Năm",     "",            "clb_khong_co","",           ""],
        ["0012345","Hoàng Văn Sáu",   "",            "clb_bongro",  "",         8.5],
        ["0012346","Vũ Thị Bảy",      "",            "clb_amnhac",  "",           ""],
        ["0012347","Đỗ Văn Tám",      "",            "clb_vanhoc",  "",           ""],
        ["12348",  "Bùi Thị Chín",    "",            "clb_bongro",  "",           ""],
    ]
    ghi_sheet(ws, ["student_id", "name", "reserve_group", "pref_1", "pref_2", "score_1"],
              dong, to_mau=VANG)
    them_sheet_ghi_chu(wb, [
        "File này CỐ Ý SAI, dùng để kiểm tra phần mềm có cảnh báo không.",
        "Nhập file TEST_01 (danh sách CLB) trước, rồi nhập file này.",
        "",
        "Sáu lỗi đã cài sẵn, và cảnh báo tương ứng phải hiện ra:",
        "",
        "1. HS202 xuất hiện HAI DÒNG (dòng 3 và 4)",
        "   -> báo: mã HS202 xuất hiện 2 lần, chỉ dòng cuối được giữ",
        "",
        "2. hs201 chỉ khác chữ hoa/thường so với HS201",
        "   -> báo: đang coi là HAI học sinh khác nhau",
        "",
        "3. HS204 ghi nhóm dự trữ 'chinh_sac' (thiếu chữ h)",
        "   -> báo: không CLB nào nhận, gợi ý 'chinh_sach'",
        "",
        "4. HS205 xếp nguyện vọng vào 'clb_khong_co'",
        "   -> báo: club không tồn tại, BỎ QUA cả học sinh này",
        "",
        "5. Mã 12348 chỉ dài 5 chữ số trong khi các mã khác dài 7",
        "   -> báo: nghi Excel đã cắt mất số 0 ở đầu",
        "",
        "6. File này có cột score_1, nhưng đây là file NGUYỆN VỌNG —",
        "   điểm chỉ nạp được từ file chọn CLB muốn thi",
        "   -> báo: điểm trong file này KHÔNG được nạp",
        "",
        "Nếu thiếu bất kỳ cảnh báo nào trong 6 cảnh báo trên, đó là lỗi",
        "của phần mềm — hãy báo lại.",
    ])
    luu(wb, "TEST_04_CO_LOI_CO_Y.xlsx")


if __name__ == "__main__":
    print("Sinh du lieu MO PHONG (seed=%d):" % SEED)
    hs = bo_sach()
    bo_co_loi()
    print("Xong: %d hoc sinh, %d CLB" % (len(hs), len(CLB)))
