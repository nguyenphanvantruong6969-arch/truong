# -*- coding: utf-8 -*-
"""Sinh BỘ VÍ DỤ dùng trong HUONG_DAN_SU_DUNG.md — 10 học sinh, 4 CLB.

    ./.venv/bin/python du_lieu_test/vi_du_huong_dan/tao_vi_du.py

⚠️  DỮ LIỆU MÔ PHỎNG. Mười cái tên dưới đây là tên bịa, không phải học
    sinh có thật. Trình bày như số liệu khảo sát thật là bịa đặt dữ liệu.

Khác hai bộ kia ở chỗ nào:

  bo_sach/   140 em — để chạy thử ở quy mô gần thật
  TEST_01..04 120 em + 1 tệp cố ý sai — để kiểm tra phần mềm có cảnh báo
  bộ này      10 em — ĐỦ NHỎ ĐỂ IN TRỌN VÀO HƯỚNG DẪN và người đọc tự
              tính lại bằng tay rồi đối chiếu với máy

Vì thế mọi con số ở đây **viết tay**, không sinh ngẫu nhiên: hướng dẫn
giải thích từng bước dựa trên đúng những con số này.

Bộ này được dựng để bốn tình huống dạy học đều xuất hiện:
  1. Có em KHÔNG được nguyện vọng 1  -> giải thích thuật toán làm gì
  2. Có em vào bằng SUẤT DỰ TRỮ      -> giải thích cột "Diện trúng tuyển"
  3. Có ĐÚNG 1 em chưa được xếp      -> giải thích tệp _chua_duoc_xep.csv
  4. Không cảnh báo nào              -> người mới không phải đoán
"""

import csv
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import Font

from tao_du_lieu_test import ghi_sheet

THU_MUC = os.path.dirname(os.path.abspath(__file__))

# (mã, tên, chỉ tiêu, chỉ tiêu dự trữ, nhóm dự trữ)
CLB = [
    ("clb_bongro",   "CLB Bóng rổ",   3, 1, "chinh_sach"),
    ("clb_tinhoc",   "CLB Tin học",   2, 0, ""),
    ("clb_mythuat",  "CLB Mỹ thuật",  2, 0, ""),
    ("clb_nauan",    "CLB Nấu ăn",    3, 0, ""),
]

# (mã, tên, nhóm dự trữ, [nguyện vọng theo thứ tự], {CLB thi: điểm})
#
# Bóng rổ 3 chỗ nhưng 6 em xếp nguyện vọng 1 -> phải có người trượt xuống.
# Một trong 3 chỗ đó là suất dự trữ, dành cho nhóm chinh_sach.
# HS10 chỉ xếp đúng một nguyện vọng vào CLB đông nhất -> không có đường lui.
HOC_SINH = [
    ("HS01", "Nguyễn Văn An",    "",           ["clb_bongro", "clb_tinhoc"],
     {"clb_bongro": 9.0, "clb_tinhoc": 7.5}),
    ("HS02", "Trần Thị Bình",    "",           ["clb_bongro", "clb_mythuat"],
     {"clb_bongro": 8.5, "clb_mythuat": 8.0}),
    ("HS03", "Lê Minh Cường",    "",           ["clb_bongro", "clb_nauan"],
     {"clb_bongro": 8.0, "clb_nauan": 7.0}),
    ("HS04", "Phạm Thu Dung",    "chinh_sach", ["clb_bongro", "clb_mythuat"],
     {"clb_bongro": 6.0, "clb_mythuat": 6.5}),
    ("HS05", "Hoàng Văn Đức",    "",           ["clb_bongro", "clb_tinhoc"],
     {"clb_bongro": 7.5, "clb_tinhoc": 9.5}),
    ("HS06", "Vũ Ngọc Giang",    "",           ["clb_tinhoc", "clb_nauan"],
     {"clb_tinhoc": 8.5, "clb_nauan": 8.0}),
    ("HS07", "Đỗ Thị Hạnh",      "",           ["clb_mythuat", "clb_nauan"],
     {"clb_mythuat": 9.0, "clb_nauan": 7.5}),
    ("HS08", "Bùi Quang Khánh",  "",           ["clb_mythuat", "clb_nauan"],
     {"clb_mythuat": 7.0, "clb_nauan": 8.5}),
    ("HS09", "Ngô Phương Linh",  "",           ["clb_nauan", "clb_tinhoc"],
     {"clb_nauan": 9.0, "clb_tinhoc": 6.5}),
    ("HS10", "Dương Bá Minh",    "",           ["clb_bongro"],
     {"clb_bongro": 7.0}),
]

SO_NV = 2
SO_THI = 2


def ghi_chu(wb, dong):
    hd = wb.create_sheet("Ghi chú")
    hd.column_dimensions["A"].width = 80
    hd.append(["DỮ LIỆU MÔ PHỎNG — KHÔNG PHẢI HỌC SINH CÓ THẬT"])
    hd["A1"].font = Font(bold=True, size=13, color="A63A2B")
    hd.append([""])
    for d in dong:
        hd.append([d])
    hd.append([""])
    hd.append(["Bộ ví dụ dùng trong HUONG_DAN_SU_DUNG.md."])
    hd.append(["Sinh bằng du_lieu_test/vi_du_huong_dan/tao_vi_du.py"])


def luu(wb, ten, header, rows):
    wb.save(os.path.join(THU_MUC, ten))
    ten_csv = ten[:-5] + ".csv"
    with io.open(os.path.join(THU_MUC, ten_csv), "w",
                 encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print("  da tao %-38s va %s" % (ten, ten_csv))


def sinh():
    # --- 1. Danh sách CLB ---
    cot = ["club_id", "name", "capacity", "reserve_capacity", "reserve_group"]
    dong = [list(c) for c in CLB]
    wb = Workbook(); ws = wb.active; ws.title = "Danh sách CLB"
    ghi_sheet(ws, cot, dong)
    ghi_chu(wb, [
        "4 câu lạc bộ, tổng 10 suất cho 10 học sinh — ĐỦ CHỖ cho tất cả.",
        "",
        "CLB Bóng rổ có 3 chỗ, trong đó 1 chỗ là SUẤT DỰ TRỮ dành cho",
        "nhóm chinh_sach. Sáu em xếp Bóng rổ làm nguyện vọng 1, nên đây",
        "là chỗ thuật toán phải làm việc.",
        "",
        "Dù đủ chỗ, vẫn sẽ có 1 em không được xếp — vì em đó chỉ xếp",
        "đúng một nguyện vọng vào CLB đông nhất. Thuật toán KHÔNG nhét",
        "học sinh vào CLB các em không chọn.",
        "",
        "NHẬP FILE NÀY TRƯỚC hai file kia.",
    ])
    luu(wb, "VIDU_01_danh_sach_CLB.xlsx", cot, dong)

    # --- 2. Chọn CLB muốn thi + điểm ---
    cot = ["student_id", "name", "reserve_group"]
    for i in range(1, SO_THI + 1):
        cot += ["test_club_%d" % i, "score_%d" % i]
    dong = []
    for ma, ten, nhom, nv, diem in HOC_SINH:
        d = [ma, ten, nhom]
        for i in range(SO_THI):
            d += [nv[i], diem[nv[i]]] if i < len(nv) else ["", ""]
        dong.append(d)
    wb = Workbook(); ws = wb.active; ws.title = "Chọn CLB muốn thi"
    ghi_sheet(ws, cot, dong)
    ghi_chu(wb, [
        "Mỗi em thi đúng những CLB mình đã xếp nguyện vọng, và điểm đã",
        "chấm sẵn nên không phải gõ tay.",
        "",
        "score_1 đi với test_club_1, score_2 đi với test_club_2.",
        "",
        "Chỉ HS04 có nhãn dự trữ (chinh_sach). Em này điểm Bóng rổ 6.0 —",
        "thấp nhất trong số em xếp Bóng rổ — nhưng vẫn vào được nhờ suất",
        "dự trữ. Đó là chỗ nhìn thấy cơ chế rõ nhất.",
    ])
    luu(wb, "VIDU_02_chon_CLB_muon_thi.xlsx", cot, dong)

    # --- 3. Nguyện vọng ---
    cot = (["student_id", "name", "reserve_group"] +
           ["pref_%d" % i for i in range(1, SO_NV + 1)])
    dong = [[ma, ten, nhom] + nv + [""] * (SO_NV - len(nv))
            for ma, ten, nhom, nv, _ in HOC_SINH]
    wb = Workbook(); ws = wb.active; ws.title = "Xếp hạng nguyện vọng"
    ghi_sheet(ws, cot, dong)
    ghi_chu(wb, [
        "pref_1 là nguyện vọng mong muốn nhất.",
        "",
        "HS10 chỉ xếp MỘT nguyện vọng, vào đúng CLB đông nhất. Em này sẽ",
        "không được xếp vào đâu — cố ý, để hướng dẫn giải thích tệp",
        "_chua_duoc_xep.csv. Ô trống ở pref_2 là bình thường.",
    ])
    luu(wb, "VIDU_03_xep_hang_nguyen_vong.xlsx", cot, dong)


if __name__ == "__main__":
    print("Sinh BO VI DU cho huong dan:")
    sinh()
    print("Xong: %d hoc sinh, %d CLB, %d suat"
          % (len(HOC_SINH), len(CLB), sum(c[2] for c in CLB)))
