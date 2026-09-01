"""Sinh BỘ SẠCH — dữ liệu mô phỏng KHÔNG có lỗi nào, để chạy thử
trọn quy trình nạp file → chấm → chạy pipeline → xuất kết quả.

    ./.venv/bin/python du_lieu_test/bo_sach/tao_bo_sach.py

⚠️  DỮ LIỆU BỊA, KHÔNG PHẢI HỌC SINH CÓ THẬT. Tên và mã do máy sinh
    (có seed nên lần nào cũng ra đúng bộ này). Trình bày như số liệu
    khảo sát thật là bịa đặt dữ liệu.

Khác bộ TEST_01..04 ở chỗ nào:

  * Thư mục này KHÔNG chứa file lỗi. Thả cả thư mục vào phần mềm cũng
    không sinh ra cảnh báo nào — đó là mục đích của bộ này.
  * Mọi lượt đăng ký thi đều CÓ ĐIỂM sẵn, nên không còn cảnh báo
    "chưa chấm điểm".
  * Mỗi em xếp 6 nguyện vọng nhưng chỉ thi 4 CLB đầu. Hai nguyện vọng
    cuối rơi vào các CLB còn nhiều chỗ, nên em nào cũng có đường lui:
    kết quả xếp được 100%, bảng xuất ra không có dòng "chưa xếp".
  * Đăng ký thi luôn là TẬP CON của nguyện vọng, nên không có lượt thi
    bỏ phí — phần mềm cảnh báo đúng ca đó, bộ sạch không được tạo ra.

Nguyện vọng 5-6 không thi vẫn được xét, nhưng ở Tier 2 (chỉ bốc thăm
STB, không có điểm). Nhờ vậy bộ này cho thấy CẢ HAI tầng ưu tiên chứ
không chỉ tầng điểm.
"""

import csv
import io
import os
import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import random

from openpyxl import Workbook
from openpyxl.styles import Font

# Dùng lại đúng bộ tên và hàm dựng sheet của bộ TEST — hai bộ trông
# giống nhau thì người chạy thử không phải học lại định dạng.
from tao_du_lieu_test import ghi_sheet, sinh_ten

THU_MUC = os.path.dirname(os.path.abspath(__file__))
SEED = 9090
SO_HOC_SINH = 140
SO_NV = 6          # mỗi em xếp 6 nguyện vọng
SO_THI = 4         # nhưng chỉ thi 4 CLB đầu — giữ số lượt thi vừa sức

# 12 CLB, tổng 150 suất cho 140 em.
#
# Ba nhóm độ hút, và điều quan trọng nhất là nhóm "còn chỗ" đủ rộng để
# đỡ hết phần tràn: 68 suất ở CLB đông + 40 suất ở CLB vừa = 108, còn
# lại 32 em phải xuống nhóm còn chỗ, mà nhóm đó có 42 suất.
CLB = [
    # (mã, tên, chỉ tiêu, chỉ tiêu dự trữ, nhóm dự trữ, độ hút)
    ("clb_bongda",     "CLB Bóng đá",      22, 5, "chinh_sach", 10.0),
    ("clb_tienganh",   "CLB Tiếng Anh",    18, 3, "khoi_10",    10.0),
    ("clb_tinhoc",     "CLB Tin học",      14, 4, "chinh_sach", 10.0),
    ("clb_mythuat",    "CLB Mỹ thuật",     14, 4, "khoi_10",    10.0),

    ("clb_bongro",     "CLB Bóng rổ",      16, 0, "",            3.0),
    ("clb_amnhac",     "CLB Âm nhạc",      14, 0, "",            3.0),
    ("clb_nhiepanh",   "CLB Nhiếp ảnh",    10, 0, "",            3.0),

    ("clb_robotics",   "CLB Robotics",     10, 0, "",            0.8),
    ("clb_khoahoc",    "CLB Khoa học",     10, 0, "",            0.8),
    ("clb_vanhoc",     "CLB Văn học",       8, 0, "",            0.8),
    ("clb_tinhnguyen", "CLB Tình nguyện",  10, 0, "",            0.8),
    ("clb_covua",      "CLB Cờ vua",        4, 0, "",            0.8),
]
CON_CHO = ["clb_robotics", "clb_khoahoc", "clb_vanhoc",
           "clb_tinhnguyen", "clb_covua"]

DIEM_TB = {"": 7.6, "chinh_sach": 6.4, "khoi_10": 6.6}
DIEM_LECH = 1.1


def chon_theo_trong_so(rng, ung_vien, trong_so, n):
    """Bốc n phần tử KHÔNG lặp, xác suất tỉ lệ với trọng số."""
    con = list(ung_vien)
    ra = []
    for _ in range(min(n, len(con))):
        tong = sum(trong_so[c] for c in con)
        moc = rng.uniform(0, tong)
        chay = 0.0
        for c in con:
            chay += trong_so[c]
            if chay >= moc:
                ra.append(c)
                con.remove(c)
                break
    return ra


def sinh_diem(rng, nhom):
    d = rng.gauss(DIEM_TB.get(nhom, DIEM_TB[""]), DIEM_LECH)
    return round(min(10.0, max(4.0, d)), 1)


def them_ghi_chu(wb, dong):
    hd = wb.create_sheet("Ghi chú")
    hd.column_dimensions["A"].width = 84
    hd.append(["DỮ LIỆU MÔ PHỎNG — KHÔNG PHẢI HỌC SINH CÓ THẬT"])
    hd["A1"].font = Font(bold=True, size=13, color="A63A2B")
    hd.append([""])
    for d in dong:
        hd.append([d])
    hd.append([""])
    hd.append(["Sinh bằng du_lieu_test/bo_sach/tao_bo_sach.py, seed = %d" % SEED])
    hd.append(["Chạy lại script luôn cho ra đúng bộ dữ liệu này."])


def luu(wb, ten, header, rows):
    """Ghi cả .xlsx và .csv.

    Máy nào thiếu thư viện đọc Excel (openpyxl) thì phần mềm bảo lưu
    sang CSV — nên bộ này kèm sẵn bản CSV, khỏi phải mở Excel ra lưu lại.
    utf-8-sig để Excel mở lên không vỡ dấu tiếng Việt.
    """
    wb.save(os.path.join(THU_MUC, ten))
    ten_csv = ten[:-5] + ".csv"
    with io.open(os.path.join(THU_MUC, ten_csv), "w",
                 encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(header)
        w.writerows(rows)
    print("  da tao %-34s va %s" % (ten, ten_csv))


def sinh():
    rng = random.Random(SEED)
    ma_clb = [c[0] for c in CLB]
    hut = {c[0]: c[5] for c in CLB}
    hut_con_cho = {c: hut[c] for c in CON_CHO}

    hoc_sinh = []
    for i in range(1, SO_HOC_SINH + 1):
        r = rng.random()
        nhom = "chinh_sach" if r < 0.13 else ("khoi_10" if r < 0.23 else "")
        hoc_sinh.append(("HS%03d" % i, sinh_ten(rng), nhom))

    # --- 1. Danh sách CLB -------------------------------------------
    wb = Workbook(); ws = wb.active; ws.title = "Danh sách CLB"
    cot_clb = ["club_id", "name", "capacity", "reserve_capacity", "reserve_group"]
    dong_clb = [list(c[:5]) for c in CLB]
    ghi_sheet(ws, cot_clb, dong_clb)
    them_ghi_chu(wb, [
        "12 câu lạc bộ, tổng 150 suất cho 140 học sinh.",
        "",
        "Bốn CLB đông nhất có suất dự trữ: Bóng đá và Tin học dành cho",
        "nhóm chinh_sach, Tiếng Anh và Mỹ thuật dành cho nhóm khoi_10.",
        "",
        "Năm CLB cuối bảng cố ý để dư chỗ. Đó là đường lui cho những em",
        "trượt hết nguyện vọng trên, và là lý do bộ này xếp được 100%.",
        "",
        "NHẬP FILE NÀY TRƯỚC hai file kia. Nếu không, mọi học sinh sẽ bị",
        "bỏ qua vì mã CLB chưa tồn tại trong phần mềm.",
    ])
    luu(wb, "SACH_01_danh_sach_CLB.xlsx", cot_clb, dong_clb)

    # --- 2 & 3. Nguyện vọng và đăng ký thi --------------------------
    dong_thi, dong_nv = [], []
    for ma, ten, nhom in hoc_sinh:
        # 4 nguyện vọng đầu bốc theo độ hút -> dồn vào CLB đông.
        dau = chon_theo_trong_so(rng, ma_clb, hut, SO_NV - 2)
        # 2 nguyện vọng cuối LUÔN nằm trong nhóm còn chỗ -> ai cũng có
        # đường lui. Bỏ những CLB đã bốc ở trên để không trùng.
        con = [c for c in CON_CHO if c not in dau]
        duoi = chon_theo_trong_so(rng, con, hut_con_cho, 2)
        nv = dau + duoi

        # Thi ĐÚNG 4 nguyện vọng đầu — tập con của nguyện vọng, nên
        # không có lượt thi bỏ phí.
        thi = nv[:SO_THI]

        o = [ma, ten, nhom]
        for i in range(SO_THI):
            o += [thi[i], sinh_diem(rng, nhom)]
        dong_thi.append(o)
        dong_nv.append([ma, ten, nhom] + nv)

    cot_thi = ["student_id", "name", "reserve_group"]
    for i in range(1, SO_THI + 1):
        cot_thi += ["test_club_%d" % i, "score_%d" % i]

    wb = Workbook(); ws = wb.active; ws.title = "Chọn CLB muốn thi"
    ghi_sheet(ws, cot_thi, dong_thi)
    them_ghi_chu(wb, [
        "140 học sinh, mỗi em đăng ký thi ĐÚNG 4 CLB và đã có điểm sẵn.",
        "",
        "score_1 đi với test_club_1, score_2 đi với test_club_2, ghép",
        "theo SỐ trong tên cột chứ không theo vị trí. Nạp file này xong",
        "là có luôn 560 ô điểm, không phải gõ tay.",
        "",
        "Vì mọi lượt thi đều có điểm nên mục Cảnh báo dữ liệu sẽ trống —",
        "không còn dòng 'CLB ... chưa chấm điểm' nào.",
        "",
        "Cột reserve_group: khoảng 23%% số em thuộc diện dự trữ",
        "(chinh_sach hoặc khoi_10), còn lại để trống.",
    ])
    luu(wb, "SACH_02_chon_CLB_muon_thi.xlsx", cot_thi, dong_thi)

    cot_nv = ["student_id", "name", "reserve_group"] + \
             ["pref_%d" % i for i in range(1, SO_NV + 1)]
    wb = Workbook(); ws = wb.active; ws.title = "Xếp hạng nguyện vọng"
    ghi_sheet(ws, cot_nv, dong_nv)
    them_ghi_chu(wb, [
        "Cùng 140 học sinh, mỗi em xếp 6 nguyện vọng theo thứ tự cột:",
        "pref_1 là mong muốn nhất.",
        "",
        "Bốn nguyện vọng đầu là những CLB em có thi (có điểm) — xét ở",
        "Tier 1. Hai nguyện vọng cuối em KHÔNG thi — vẫn được xét, nhưng",
        "ở Tier 2, chỉ dựa vào số bốc thăm STB.",
        "",
        "Nhờ hai tầng đó mà bảng kết quả cho thấy cả hai cơ chế, và mọi",
        "em đều có chỗ.",
    ])
    luu(wb, "SACH_03_xep_hang_nguyen_vong.xlsx", cot_nv, dong_nv)
    return hoc_sinh


if __name__ == "__main__":
    print("Sinh BO SACH (seed=%d):" % SEED)
    hs = sinh()
    print("Xong: %d hoc sinh, %d CLB, %d suat"
          % (len(hs), len(CLB), sum(c[2] for c in CLB)))
