"""Sinh các file mẫu .xlsx từ chính các file .csv mẫu trong thư mục này.

Chạy lại mỗi khi sửa file .csv mẫu, để hai bộ không lệch nhau:

    ./.venv/bin/python mau_csv/tao_mau_excel.py

Bố cục mỗi file: SHEET ĐẦU là dữ liệu (phần mềm chỉ đọc sheet đầu), sheet
sau là hướng dẫn điền cho giáo viên. Nhờ vậy ghi chú không lẫn vào dữ liệu.
"""

import csv
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

THU_MUC = os.path.dirname(os.path.abspath(__file__))

# (file csv nguồn, tên file xlsx, tiêu đề, các dòng hướng dẫn)
BO_MAU = [
    (
        "05_danh_sach_club.csv",
        "MAU_01_danh_sach_CLB.xlsx",
        "Danh sách CLB",
        [
            "Điền mỗi CLB một dòng. Nhập file này TRƯỚC hai file học sinh.",
            "",
            "club_id           BẮT BUỘC. Mã CLB, phải khớp từng ký tự với mã",
            "                  dùng trong hai file học sinh.",
            "name              BẮT BUỘC. Tên đầy đủ, hiện trên màn hình và in",
            "                  ra file kết quả.",
            "capacity          BẮT BUỘC. Tổng chỉ tiêu, phải lớn hơn 0.",
            "reserve_capacity  Tuỳ chọn. Số suất dành cho nhóm dự trữ.",
            "                  Bỏ trống = 0. Không được lớn hơn capacity.",
            "reserve_group     Tuỳ chọn. Tên nhóm được ưu tiên, ví dụ",
            "                  chinh_sach. Bỏ trống = CLB không có dự trữ.",
            "",
            "Nhập lại file này là CẬP NHẬT CLB đã có theo club_id, không tạo",
            "trùng. Dòng nào chỉ tiêu sai thì bỏ riêng dòng đó kèm cảnh báo.",
        ],
    ),
    (
        "01_chon_club_thi_dang_rong.csv",
        "MAU_02_chon_CLB_muon_thi.xlsx",
        "Chọn CLB muốn thi",
        [
            "Mỗi học sinh MỘT dòng. Đây là bước tick chọn, KHÔNG xếp thứ tự.",
            "",
            "student_id     BẮT BUỘC. Mã học sinh, phải duy nhất và giống hệt",
            "               mã dùng ở file nguyện vọng.",
            "name           Tuỳ chọn. Chỉ dùng khi tạo mới học sinh.",
            "reserve_group  Tuỳ chọn. Nhóm dự trữ của HỌC SINH, ví dụ",
            "               chinh_sach. Ô trống thì giữ nguyên nhóm đã có,",
            "               không bị xoá.",
            "test_club_1…   Mã CLB muốn dự tuyển. Ô trống được bỏ qua, không",
            "               cần điền kín. Thêm cột test_club_5, 6… nếu cần.",
            "",
            "Mọi mã CLB phải đã tồn tại. Học sinh có bất kỳ mã sai nào sẽ bị",
            "bỏ qua TOÀN BỘ kèm cảnh báo — phần mềm không nhập một nửa.",
        ],
    ),
    (
        "03_nguyen_vong_dang_rong.csv",
        "MAU_03_xep_hang_nguyen_vong.xlsx",
        "Xếp hạng nguyện vọng",
        [
            "Mỗi học sinh MỘT dòng. THỨ TỰ CỘT chính là thứ tự nguyện vọng.",
            "",
            "student_id     BẮT BUỘC. Mã học sinh.",
            "name           Tuỳ chọn. Chỉ dùng khi tạo mới học sinh.",
            "reserve_group  Tuỳ chọn. Nhóm dự trữ của HỌC SINH. Ô trống thì",
            "               giữ nguyên nhóm đã có, không bị xoá.",
            "pref_1         Nguyện vọng 1 — mong muốn nhất.",
            "pref_2, pref_3 Các nguyện vọng tiếp theo. Tối đa 10 nguyện vọng.",
            "",
            "Học sinh có quá 10 nguyện vọng sẽ bị bỏ qua TOÀN BỘ — phần mềm",
            "không tự cắt còn 10, vì cắt bớt là âm thầm đổi nguyện vọng của",
            "học sinh. CLB trùng nhau chỉ giữ lần đầu, kèm cảnh báo.",
        ],
    ),
]

XANH = PatternFill("solid", fgColor="DCE9E0")


# Điền sẵn nhóm cho vài em trong mẫu. Để trống hết thì giáo viên mở ra
# thấy một cột rỗng và không đoán được phải viết gì vào đó — mà bỏ trống
# cả cột thì cơ chế dự trữ của RB-DA vô hiệu hoàn toàn, pipeline vẫn chạy
# và không báo lỗi gì. Giá trị này khớp với reserve_group khai trong file
# danh sách CLB mẫu; hai bên PHẢI khớp từng ký tự.
NHOM_MAU = {"HS002": "chinh_sach", "HS005": "chinh_sach"}


def them_cot_reserve_group(header, rows):
    """File CSV mẫu chưa có cột reserve_group; bản Excel thì có, vì đây
    là chỗ giáo viên điền tay và rất dễ quên nếu không thấy cột."""
    vi_tri = 2 if len(header) > 2 else len(header)
    header = header[:vi_tri] + ["reserve_group"] + header[vi_tri:]
    rows = [
        r[:vi_tri] + [NHOM_MAU.get(r[0] if r else "", "")] + r[vi_tri:]
        for r in rows
    ]
    return header, rows


def tao_mot_file(ten_csv, ten_xlsx, tieu_de, huong_dan):
    with open(os.path.join(THU_MUC, ten_csv), encoding="utf-8-sig", newline="") as f:
        bang = list(csv.reader(f))
    header, rows = bang[0], bang[1:]

    if "student_id" in header and "reserve_group" not in header:
        header, rows = them_cot_reserve_group(header, rows)

    wb = Workbook()
    ws = wb.active
    ws.title = tieu_de[:31]          # Excel giới hạn 31 ký tự cho tên sheet
    ws.append(header)
    for r in rows:
        ws.append(r)

    for i, ten_cot in enumerate(header, start=1):
        o = ws.cell(row=1, column=i)
        o.font = Font(bold=True)
        o.fill = XANH
        o.alignment = Alignment(vertical="center")
        rong = max([len(ten_cot)] + [len(str(r[i - 1])) for r in rows if i - 1 < len(r)])
        ws.column_dimensions[get_column_letter(i)].width = min(max(rong + 4, 12), 30)
    ws.freeze_panes = "A2"

    # Sheet hướng dẫn nằm SAU sheet dữ liệu — phần mềm chỉ đọc sheet đầu
    # nên ghi chú ở đây không bao giờ lẫn vào dữ liệu.
    hd = wb.create_sheet("Hướng dẫn điền")
    hd.column_dimensions["A"].width = 78
    hd.append([tieu_de])
    hd["A1"].font = Font(bold=True, size=13)
    hd.append([""])
    for dong in huong_dan:
        hd.append([dong])
    hd.append([""])
    hd.append(["Xem đầy đủ: mau_csv/HUONG_DAN_CSV.md"])

    duong_dan = os.path.join(THU_MUC, ten_xlsx)
    wb.save(duong_dan)
    return duong_dan, len(rows)


if __name__ == "__main__":
    for ten_csv, ten_xlsx, tieu_de, huong_dan in BO_MAU:
        p, n = tao_mot_file(ten_csv, ten_xlsx, tieu_de, huong_dan)
        print("da tao %s (%d dong du lieu)" % (os.path.basename(p), n))
