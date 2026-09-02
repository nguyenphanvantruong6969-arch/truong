"""
api.py
======
Lớp API expose cho pywebview — mọi hàm public ở đây có thể được gọi
từ JS qua window.pywebview.api.<ten_ham>(...) và luôn trả về dict
(pywebview tự serialize JSON hai chiều).

Quy ước trả về thống nhất cho mọi hàm:
    { "ok": bool, "data": ..., "errors": [...] }
Để JS chỉ cần check `.ok` là biết thành công hay không, không phải
try/catch riêng lẻ từng hàm.

Mỗi phần tử trong "errors" (và trong các danh sách "warnings" trả về
kèm data khi ok=True) là MỘT trong hai dạng:
  - chuỗi thô (vd traceback) — hiển thị nguyên văn, không dịch.
  - {"code": "...", "params": {...}} — do err() ở i18n_errors.py tạo ra,
    frontend tự dịch sang ngôn ngữ đang chọn (xem ERROR_MESSAGES trong
    app.js, phải khớp với MESSAGES trong i18n_errors.py).
Điều này cho phép app hỗ trợ song ngữ (vi/en) mà không cần Python biết
người dùng đang chọn ngôn ngữ nào.
"""

import csv
import datetime
import io
import os
import sys
import traceback

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from rbda_priority_pipeline import (
    init_db,
    load_from_sqlite,
    validate_data_integrity,
    generate_stb_lottery,
    run_rbda,
    sanity_check_result,
    verify_stability,
    export_match_results,
    default_reserve_eligible_fn,
    connect_db,
)
from i18n_errors import err

import sqlite3
import statistics


def _now() -> str:
    return datetime.datetime.now().isoformat(timespec="seconds")


def _ok(data=None):
    return {"ok": True, "data": data, "errors": []}


def _fail(errors):
    if isinstance(errors, str):
        errors = [errors]
    elif isinstance(errors, dict) and "code" in errors and "params" in errors:
        errors = [errors]
    return {"ok": False, "data": None, "errors": errors}


class PipelineAPI:
    def __init__(self, db_path: str):
        self.db_path = db_path
        # DẤU GẠCH DƯỚI LÀ BẮT BUỘC, không phải quy ước cho đẹp. Xem
        # _set_window() ngay dưới đây.
        self._window = None
        init_db(self.db_path)

    def _set_window(self, window) -> None:
        """
        Gọi từ main.py sau webview.create_window(...):
            api = PipelineAPI(db_path)
            window = webview.create_window("...", "index.html", js_api=api)
            api._set_window(window)
        Không bắt buộc — nếu không gọi, mọi tính năng vẫn hoạt động,
        chỉ riêng import CSV sẽ nhận nội dung file qua FileReader ở JS
        (đã dùng theo mặc định) thay vì hộp thoại chọn file gốc của hệ điều hành.

        VÌ SAO TÊN PHẢI BẮT ĐẦU BẰNG DẤU GẠCH DƯỚI — đây từng là lỗi làm
        TREO HẲN app trên Windows (cửa sổ ghi "Not Responding"):

        pywebview dựng cầu nối bằng cách DÒ chính đối tượng API này
        (webview/util.py, get_functions). Luật của nó là:

            for name in dir(obj):
                if name.startswith('_'): continue      # <- lối thoát duy nhất
                attr = getattr(obj, name)              # <- KÍCH HOẠT property
                ...
                elif not callable(attr) and hasattr(attr, '__module__'):
                    get_functions(attr, ...)           # <- ĐỆ QUY vào attr

        Cửa sổ pywebview không callable và có `__module__`, nên khi nó nằm ở
        một thuộc tính CÔNG KHAI, pywebview đệ quy vào chính cửa sổ của nó —
        và `dir()` cửa sổ đó có bốn property CHẶN (webview/window.py):

            @property
            def width(self):
                self.events.shown.wait(15)              # chờ tới 15 giây
                width, _ = self.gui.get_size(self.uid)  # đọc Control.Size

        `width`, `height`, `x`, `y`: mỗi cái chờ tới 15 giây, và `get_size`
        đọc thuộc tính của một control WinForms TỪ LUỒNG KHÁC luồng giao
        diện. Luồng dò bị chặn -> finish.js không bao giờ chạy ->
        `window.pywebview.api` mãi rỗng -> giao diện báo không kết nối được,
        còn cửa sổ thì treo.

        Dấu gạch dưới cắt đứt chuỗi đó ngay bước đầu: `get_functions` bỏ qua
        tên bắt đầu bằng `_` TRƯỚC khi gọi `getattr`, nên cửa sổ không hề bị
        chạm tới. Có test canh (tests/test_do_api.py).
        """
        self._window = window

    # -----------------------------------------------------------------
    # TAB "VẬN HÀNH PIPELINE"
    # -----------------------------------------------------------------

    def get_last_run_info(self):
        """Thông tin lần chạy pipeline gần nhất — để hiện ở sidebar/dashboard."""
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            row = cur.execute("SELECT * FROM run_meta WHERE id = 1").fetchone()
            conn.close()
            return _ok(dict(row) if row else None)
        except Exception as e:
            return _fail(err("error_reading_last_run", detail=str(e)))

    def get_dashboard_status(self):
        """Số liệu tổng quan để hiển thị ngay khi mở tab Pipeline."""
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            n_students = cur.execute("SELECT COUNT(*) FROM students").fetchone()[0]
            n_clubs = cur.execute("SELECT COUNT(*) FROM clubs").fetchone()[0]
            n_prefs = cur.execute(
                "SELECT COUNT(DISTINCT student_id) FROM preferences"
            ).fetchone()[0]
            n_matched = cur.execute(
                "SELECT COUNT(*) FROM match_results WHERE club_id IS NOT NULL"
            ).fetchone()[0]
            has_results = cur.execute(
                "SELECT COUNT(*) FROM match_results"
            ).fetchone()[0] > 0
            conn.close()
            return _ok({
                "n_students": n_students,
                "n_clubs": n_clubs,
                "n_students_with_preferences": n_prefs,
                "n_matched": n_matched,
                "has_results": has_results,
            })
        except Exception as e:
            return _fail(err("error_reading_dashboard", detail=str(e)))

    def check_data_integrity(self):
        """Nút 'Kiểm tra dữ liệu' — chỉ validate, KHÔNG chạy pipeline."""
        try:
            students, clubs, tested_scores, applicants, preferences, _ = (
                load_from_sqlite(self.db_path)
            )
            errors = validate_data_integrity(students, clubs, preferences, applicants)
            if errors:
                return _fail(errors)
            return _ok({
                "message": "Du lieu hop le.",
                "n_students": len(students),
                "n_clubs": len(clubs),
            })
        except Exception as e:
            return _fail([err("error_checking_integrity", detail=str(e)), traceback.format_exc()])

    # -----------------------------------------------------------------
    # KIỂM TRA SỨC KHOẺ DỮ LIỆU (pre-flight)
    #
    # validate_data_integrity() chỉ bắt dữ liệu KHÔNG HỢP LỆ (club không
    # tồn tại, capacity <= 0, nguyện vọng trùng…). Nhưng có cả một nhóm
    # tình huống mà dữ liệu VẪN HỢP LỆ, pipeline VẪN CHẠY, kết quả VẪN
    # trông bình thường — trong khi ai được vào club đã bị thay đổi bởi
    # một thiếu sót mà người vận hành không hề thấy. Ví dụ nguy hiểm nhất:
    # giáo viên mới chấm được một nửa danh sách, nửa còn lại lập tức bị
    # xếp dưới TOÀN BỘ các em đã có điểm — kể cả em thấp điểm nhất.
    #
    # Hàm này liệt kê đúng nhóm đó dưới dạng CẢNH BÁO (không phải lỗi):
    # không chặn chạy pipeline, nhưng bắt buộc phải hiện ra trước mắt
    # người vận hành để họ tự quyết định.
    # -----------------------------------------------------------------

    _HEALTH_SAMPLE_LIMIT = 5

    # Nguong bao "diem la": gap bao nhieu lan TRUNG VI cua chinh CLB do.
    #
    # Chon 3.0 bang so do, khong phai cam tinh. Do tren 579 o diem that
    # cua hai bo du lieu: ti le (diem cao nhat / trung vi) trong mot CLB
    # cao nhat la 1.42. Loi go lech dau cham thi lech gap ~10 lan. Giua
    # 1.42 va 10 la mot khoang trong gan mot bac do lon.
    #
    # Quet thu: he so 1.5 -> 11 canh bao GIA; tu 2.0 tro len -> 0. Lay
    # 3.0 de cach xa ca that te nhat 2.1 lan ma van thua suc bat loi go.
    _NGUONG_DIEM_LA = 3.0

    def get_data_health_report(self):
        """
        Rà soát các thiếu sót dữ liệu ÂM THẦM LÀM ĐỔI KẾT QUẢ.
        Trả về {"warnings": [...], "n_warnings": int, "n_high": int} —
        mỗi cảnh báo là {code, params, severity} với severity là
        "high" (gần như chắc chắn làm sai kết quả) hoặc "medium"/"info".
        """
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            warnings = []

            def warn(severity, code, **params):
                entry = err(code, **params)
                entry["severity"] = severity
                warnings.append(entry)

            # --- 1. Chấm điểm thiếu / chưa chấm ---------------------------
            for row in cur.execute("""
                SELECT c.club_id,
                       COUNT(DISTINCT t.student_id) AS n_applicants,
                       COUNT(DISTINCT sc.student_id) AS n_scored
                FROM clubs c
                JOIN club_test_selection t ON t.club_id = c.club_id
                LEFT JOIN club_scores sc
                       ON sc.club_id = c.club_id AND sc.student_id = t.student_id
                GROUP BY c.club_id
                ORDER BY c.club_id
            """).fetchall():
                n_app, n_scored = row["n_applicants"], row["n_scored"]
                if n_app == 0:
                    continue
                if n_scored == 0:
                    warn("high", "health_scoring_none",
                         club_id=row["club_id"], n_applicants=n_app)
                elif n_scored < n_app:
                    warn("high", "health_scoring_partial",
                         club_id=row["club_id"], n_applicants=n_app,
                         n_scored=n_scored, n_missing=n_app - n_scored)

            # --- 2. Đăng ký thi nhưng không xếp nguyện vọng club đó -------
            wasted = cur.execute("""
                SELECT t.student_id, t.club_id
                FROM club_test_selection t
                LEFT JOIN preferences p
                       ON p.student_id = t.student_id AND p.club_id = t.club_id
                WHERE p.student_id IS NULL
                ORDER BY t.student_id, t.club_id
            """).fetchall()
            if wasted:
                sample = ", ".join(
                    f"{r['student_id']}→{r['club_id']}"
                    for r in wasted[:self._HEALTH_SAMPLE_LIMIT]
                )
                warn("high", "health_tested_not_ranked",
                     n=len(wasted), sample=sample)

            # --- 3. Học sinh chưa xếp nguyện vọng nào ---------------------
            no_pref = cur.execute("""
                SELECT s.student_id FROM students s
                LEFT JOIN preferences p ON p.student_id = s.student_id
                WHERE p.student_id IS NULL
                ORDER BY s.student_id
            """).fetchall()
            if no_pref:
                warn("medium", "health_student_no_preferences",
                     n=len(no_pref),
                     sample=", ".join(r["student_id"] for r in no_pref[:self._HEALTH_SAMPLE_LIMIT]))

            # --- 4. Nhãn dự trữ của học sinh mà không club nào dùng -------
            # Kèm MÃ HỌC SINH như hai mục trên. Không có mã thì cảnh báo
            # bảo "kiểm tra xem có gõ sai chính tả không" mà người dùng
            # phải tự dò tay qua cả danh sách mới biết dò em nào.
            for row in cur.execute("""
                SELECT s.reserve_group AS g, COUNT(*) AS n,
                       GROUP_CONCAT(s.student_id) AS ids
                FROM students s
                WHERE s.reserve_group IS NOT NULL AND TRIM(s.reserve_group) <> ''
                  AND s.reserve_group NOT IN (
                      SELECT reserve_group FROM clubs
                      WHERE reserve_group IS NOT NULL AND TRIM(reserve_group) <> ''
                  )
                GROUP BY s.reserve_group ORDER BY s.reserve_group
            """).fetchall():
                # GROUP_CONCAT không hứa thứ tự — phải sắp TRƯỚC khi cắt,
                # nếu không mỗi lần bấm "Kiểm tra lại" lại ra một mẫu khác
                # và người dùng tưởng dữ liệu vừa đổi.
                ids = sorted((row["ids"] or "").split(","))
                warn("high", "health_orphan_student_group",
                     reserve_group=row["g"], n=row["n"],
                     sample=", ".join(ids[:self._HEALTH_SAMPLE_LIMIT]))

            # --- 5. Club có suất dự trữ nhưng chưa đặt nhãn ---------------
            for row in cur.execute("""
                SELECT club_id, reserve_capacity FROM clubs
                WHERE reserve_capacity > 0
                  AND (reserve_group IS NULL OR TRIM(reserve_group) = '')
                ORDER BY club_id
            """).fetchall():
                warn("high", "health_club_reserve_no_group",
                     club_id=row["club_id"], reserve_capacity=row["reserve_capacity"])

            # --- 6. Club dành suất cho nhãn chưa học sinh nào mang --------
            for row in cur.execute("""
                SELECT club_id, reserve_group, reserve_capacity FROM clubs
                WHERE reserve_capacity > 0
                  AND reserve_group IS NOT NULL AND TRIM(reserve_group) <> ''
                  AND reserve_group NOT IN (
                      SELECT reserve_group FROM students
                      WHERE reserve_group IS NOT NULL AND TRIM(reserve_group) <> ''
                  )
                ORDER BY club_id
            """).fetchall():
                warn("medium", "health_club_group_no_students",
                     club_id=row["club_id"], reserve_group=row["reserve_group"],
                     reserve_capacity=row["reserve_capacity"])

            # --- 7. Tổng chỗ ít hơn số học sinh đã nộp nguyện vọng --------
            n_seats = cur.execute(
                "SELECT COALESCE(SUM(capacity), 0) FROM clubs"
            ).fetchone()[0]
            n_with_prefs = cur.execute(
                "SELECT COUNT(DISTINCT student_id) FROM preferences"
            ).fetchone()[0]
            if n_with_prefs > n_seats:
                warn("info", "health_oversubscribed",
                     n_seats=n_seats, n_students=n_with_prefs,
                     n_short=n_with_prefs - n_seats)

            # --- 8. Điểm lệch hẳn khỏi phân bố của chính CLB đó ----------
            # Gõ 70 thay vì 7.0 thì phần mềm vẫn nhận, và em đó nhảy lên
            # đầu bảng. Đo trên bộ ví dụ: MỘT lỗi gõ làm BA em đổi chỗ,
            # vì em bị đẩy ra lại đi đẩy em khác.
            #
            # KHÔNG đặt trần cứng ở 10 — trường có thể chấm thang 100, và
            # chặn cứng là chặn nhầm. So với TRUNG VỊ CỦA CHÍNH CLB đó thì
            # thang nào cũng đúng, và bắt được cả hai phía: 70 (thừa) lẫn
            # 0.85 (thiếu).
            diem_theo_clb: dict = {}
            for row in cur.execute("""
                SELECT club_id, student_id, score FROM club_scores
                ORDER BY club_id, student_id
            """).fetchall():
                diem_theo_clb.setdefault(row["club_id"], []).append(
                    (row["student_id"], row["score"])
                )
            for club_id in sorted(diem_theo_clb):
                cap = diem_theo_clb[club_id]
                # Dưới 3 điểm thì không có phân bố nào để mà so.
                if len(cap) < 3:
                    continue
                giua = statistics.median([s for _, s in cap])
                if giua <= 0:
                    continue
                la = [(sid, s) for sid, s in cap
                      if s > self._NGUONG_DIEM_LA * giua
                      or s * self._NGUONG_DIEM_LA < giua]
                if la:
                    warn("high", "health_score_outlier",
                         club_id=club_id, n=len(la),
                         trung_vi=("%g" % giua),
                         sample=", ".join(
                             "%s (%g)" % (sid, s)
                             for sid, s in la[:self._HEALTH_SAMPLE_LIMIT]))

            conn.close()
            return _ok({
                "warnings": warnings,
                "n_warnings": len(warnings),
                "n_high": sum(1 for w in warnings if w["severity"] == "high"),
            })
        except Exception as e:
            return _fail(err("error_checking_integrity", detail=str(e)))

    def get_pipeline_run_warning(self):
        """
        Gọi TRƯỚC khi hiện hộp xác nhận 'Chạy pipeline' (bước 1 trong xác
        nhận 2 bước). Cho UI biết: (a) đã có kết quả cũ sẽ bị ghi đè chưa,
        (b) STB đã bị khoá chưa (nếu khoá, chạy bình thường sẽ TÁI SỬ DỤNG
        STB cũ chứ không vẽ lại — chỉ vẽ lại nếu force_redraw_stb=True).
        """
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            has_results = cur.execute(
                "SELECT COUNT(*) FROM match_results"
            ).fetchone()[0] > 0
            lock_row = cur.execute(
                "SELECT is_locked, locked_at FROM stb_lock WHERE id = 1"
            ).fetchone()
            last_run = cur.execute(
                "SELECT run_at, seed FROM run_meta WHERE id = 1"
            ).fetchone()
            conn.close()
            return _ok({
                "has_existing_results": has_results,
                "stb_locked": bool(lock_row[0]) if lock_row else False,
                "stb_locked_at": lock_row[1] if lock_row else None,
                "last_run_at": last_run[0] if last_run else None,
                "last_run_seed": last_run[1] if last_run else None,
            })
        except Exception as e:
            return _fail(err("error_checking_run_warning", detail=str(e)))

    def get_stb_lock_status(self):
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            row = cur.execute(
                "SELECT is_locked, locked_at, unlocked_at FROM stb_lock WHERE id = 1"
            ).fetchone()
            conn.close()
            if not row:
                return _ok({"is_locked": False, "locked_at": None, "unlocked_at": None})
            return _ok({
                "is_locked": bool(row[0]),
                "locked_at": row[1],
                "unlocked_at": row[2],
            })
        except Exception as e:
            return _fail(err("error_reading_stb_lock", detail=str(e)))

    def get_run_history(self, limit: int = 20):
        """Nhật ký toàn bộ các lần chạy pipeline, mới nhất trước — phục vụ kiểm toán."""
        try:
            limit = max(1, min(int(limit), 200))
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            rows = cur.execute(
                "SELECT run_id, seed, run_at, rounds_run, n_matched, n_total, stb_redrawn "
                "FROM run_history ORDER BY run_id DESC LIMIT ?",
                (limit,),
            ).fetchall()
            conn.close()
            return _ok([dict(r) for r in rows])
        except Exception as e:
            return _fail(err("error_reading_run_history", detail=str(e)))

    _MAX_BACKUPS = 10

    def _backup_db(self) -> str:
        """
        Sao lưu app.db bằng SQLite Backup API (connection.backup(), KHÔNG
        phải copy file thô) ngay TRƯỚC khi pipeline bắt đầu ghi gì —
        an toàn kể cả khi có tiến trình khác đang mở file cùng lúc, khác
        với copy file trực tiếp vốn có thể chụp phải trạng thái nửa-ghi
        (xem ke-hoach-mat-du-lieu.html). Giữ lại tối đa _MAX_BACKUPS bản
        gần nhất, tự xoá bản cũ hơn. Trả về đường dẫn bản sao lưu.
        """
        backup_dir = os.path.dirname(self.db_path) or "."
        ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
        backup_name = f"{os.path.basename(self.db_path)}.bak-{ts}"
        backup_path = os.path.join(backup_dir, backup_name)

        src = connect_db(self.db_path)
        try:
            dst = sqlite3.connect(backup_path)
            try:
                src.backup(dst)
            finally:
                dst.close()
        finally:
            src.close()

        prefix = f"{os.path.basename(self.db_path)}.bak-"
        existing = sorted(f for f in os.listdir(backup_dir) if f.startswith(prefix))
        for stale in existing[: -self._MAX_BACKUPS] if len(existing) > self._MAX_BACKUPS else []:
            try:
                os.remove(os.path.join(backup_dir, stale))
            except OSError:
                pass
        return backup_path

    def run_pipeline(self, seed: int = 42, force_redraw_stb: bool = False):
        """
        Nút 'Chạy pipeline' — chạy trọn 5 bước, trả về log từng bước
        để UI hiển thị lên stepper theo thời gian thực (từng bước một,
        không phải chỉ kết quả cuối).

        Quy tắc khoá STB (giải quyết #3 ghi đè âm thầm + #4 khoá STB):
          - Nếu STB CHƯA khoá (lần chạy đầu tiên, hoặc sau khi người
            dùng chủ động mở khoá): vẽ STB cho MỌI học sinh, rồi tự
            động khoá lại ngay sau khi vẽ xong.
          - Nếu STB ĐÃ khoá và force_redraw_stb=False (mặc định): KHÔNG
            vẽ lại — chỉ vẽ bổ sung cho học sinh MỚI (stb_number NULL,
            vd học sinh vừa được thêm qua kiosk sau khi đã khoá), số đã
            có giữ nguyên. Đây là hành vi mặc định khi bấm 'Chạy lại'.
          - Nếu STB đã khoá và force_redraw_stb=True (người dùng xác
            nhận 2 bước trên UI để "vẽ lại"): vẽ lại TOÀN BỘ, ghi log
            unlocked_at của lần khoá cũ rồi khoá lại với thời điểm mới.
          - Mọi lần chạy (kể cả tái sử dụng STB) đều được append vào
            run_history — không bao giờ mất dấu vết lần chạy nào.

        Tính nguyên tử (giải quyết #1 trong ke-hoach-mat-du-lieu.html):
          - Vẽ/khoá STB, ghi match_results, ghi run_meta/run_history đều
            diễn ra trong MỘT connection/transaction duy nhất. Nếu bất kỳ
            bước nào ở giữa ném lỗi (kể cả crash tiến trình được bắt
            bằng try/except ở đây), TOÀN BỘ transaction rollback — kể cả
            số STB vừa vẽ ("full rollback", phương án đã chốt: mọi lần
            gọi lại run_pipeline() sau một crash bắt đầu từ trạng thái
            sạch, không có "khoá STB mồ côi" không đi kèm kết quả nào).
          - Xuất CSV chỉ thực hiện SAU KHI transaction đã commit thành
            công — CSV là sản phẩm phụ, lỗi ghi file không được phép
            khiến DB rơi vào trạng thái dở dang.
        """
        steps_log = []
        conn = None
        try:
            try:
                backup_path = self._backup_db()
                steps_log.append({
                    "step": "backup", "status": "done",
                    "detail": err("db_backed_up", backup_name=os.path.basename(backup_path)),
                })
            except Exception as e:
                # Sao luu la luoi an toan, KHONG phai dieu kien tien quyet —
                # loi sao luu khong duoc chan pipeline chay.
                steps_log.append({
                    "step": "backup", "status": "error",
                    "detail": err("db_backup_failed", detail=str(e)),
                })

            steps_log.append({"step": "validate", "status": "running"})
            students, clubs, tested_scores, applicants, preferences, _ = (
                load_from_sqlite(self.db_path)
            )
            errors = validate_data_integrity(students, clubs, preferences, applicants)
            if errors:
                steps_log.append({"step": "validate", "status": "error", "detail": errors})
                return _fail({"steps": steps_log, "errors": errors})
            steps_log.append({"step": "validate", "status": "done"})

            steps_log.append({"step": "stb_lottery", "status": "running"})
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            lock_row = cur.execute(
                "SELECT is_locked FROM stb_lock WHERE id = 1"
            ).fetchone()
            already_locked = bool(lock_row[0]) if lock_row else False
            stb_redrawn = False

            if not already_locked or force_redraw_stb:
                # Ve lai TOAN BO
                stb_lottery_new = generate_stb_lottery(list(students.keys()), seed=seed)
                conn.executemany(
                    "UPDATE students SET stb_number = ? WHERE student_id = ?",
                    [(v, k) for k, v in stb_lottery_new.items()],
                )
                if already_locked and force_redraw_stb:
                    cur.execute(
                        "UPDATE stb_lock SET unlocked_at = ? WHERE id = 1", (_now(),)
                    )
                cur.execute(
                    "UPDATE stb_lock SET is_locked = 1, locked_at = ? WHERE id = 1",
                    (_now(),),
                )
                stb_redrawn = True
                stb_step_detail = err("stb_redrawn_and_locked", n=len(stb_lottery_new))
                for sid, v in stb_lottery_new.items():
                    students[sid]["stb"] = v
            else:
                # Da khoa: chi ve bo sung cho hoc sinh moi (stb_number con NULL)
                missing = [
                    sid for sid, info in students.items() if info.get("stb") is None
                ]
                if missing:
                    supplement = generate_stb_lottery(missing, seed=seed)
                    # danh so tiep noi sau STB lon nhat hien co, tranh trung
                    existing_max_row = cur.execute(
                        "SELECT MAX(stb_number) FROM students"
                    ).fetchone()
                    offset = (existing_max_row[0] + 1) if existing_max_row[0] is not None else 0
                    conn.executemany(
                        "UPDATE students SET stb_number = ? WHERE student_id = ?",
                        [(v + offset, k) for k, v in supplement.items()],
                    )
                    stb_step_detail = err("stb_supplemented", n=len(missing))
                    for sid, v in supplement.items():
                        students[sid]["stb"] = v + offset
                else:
                    stb_step_detail = err("stb_reused")
            steps_log.append({"step": "stb_lottery", "status": "done", "detail": stb_step_detail})

            # Khong mo connection moi de doc lai stb_number: cac thay doi
            # o tren CHUA commit, mot connection khac se khong thay duoc
            # (va se pha vo tinh nguyen tu). Cap nhat truc tiep vao dict
            # students trong bo nho (da lam o hai nhanh phia tren) roi
            # dung lai cho RB-DA — tuong duong voi doc lai tu DB nhung
            # van nam trong CUNG MOT transaction.
            stb_lottery = {sid: info["stb"] for sid, info in students.items()}

            steps_log.append({"step": "rbda_cascade", "status": "running"})
            reserve_fn = default_reserve_eligible_fn(students, clubs)
            result = run_rbda(
                students, clubs, tested_scores, applicants, preferences,
                stb_lottery, is_reserve_eligible_fn=reserve_fn,
            )
            sanity_problems = sanity_check_result(result, clubs, preferences)
            stability_problems = verify_stability(result, clubs, preferences, reserve_fn)
            if sanity_problems or stability_problems:
                steps_log.append({
                    "step": "rbda_cascade", "status": "error",
                    "detail": sanity_problems + stability_problems,
                })
                conn.rollback()
                conn.close()
                conn = None
                steps_log.append({"step": "rollback", "status": "done", "detail": err("pipeline_rolled_back")})
                return _fail({"steps": steps_log, "errors": sanity_problems + stability_problems})
            steps_log.append({
                "step": "rbda_cascade", "status": "done",
                "detail": err("rbda_done", rounds=result.rounds_run),
            })

            steps_log.append({"step": "write_results", "status": "running"})
            cur.execute("DELETE FROM match_results")
            cur.executemany(
                "INSERT INTO match_results (student_id, club_id, round_num, matched_tier, rank_in_student_pref) "
                "VALUES (?, ?, ?, ?, ?)",
                [
                    (
                        sid, cid, result.rounds_run,
                        result.matched_tier.get(sid),
                        result.rank_in_student_pref.get(sid),
                    )
                    for sid, cid in result.assignment.items()
                ],
            )
            steps_log.append({"step": "write_results", "status": "done"})

            n_matched = sum(1 for v in result.assignment.values() if v)
            run_at = _now()

            cur.execute(
                "INSERT INTO run_meta (id, seed, run_at, rounds_run, n_matched, n_total) "
                "VALUES (1, ?, ?, ?, ?, ?) "
                "ON CONFLICT(id) DO UPDATE SET seed=excluded.seed, run_at=excluded.run_at, "
                "rounds_run=excluded.rounds_run, n_matched=excluded.n_matched, n_total=excluded.n_total",
                (seed, run_at, result.rounds_run, n_matched, len(result.assignment)),
            )
            # run_history: KHONG BAO GIO ghi de — moi lan chay them 1 dong moi,
            # de nguoi dung luon xem lai duoc lich su chay pipeline (giai quyet #3).
            cur.execute(
                "INSERT INTO run_history (seed, run_at, rounds_run, n_matched, n_total, stb_redrawn) "
                "VALUES (?, ?, ?, ?, ?, ?)",
                (seed, run_at, result.rounds_run, n_matched, len(result.assignment),
                 1 if stb_redrawn else 0),
            )

            # Diem commit DUY NHAT cua toan bo pipeline: neu bat ky dong
            # nao o tren (ve STB, ghi ket qua, ghi run_meta/run_history)
            # nem exception, khoi except ben duoi se rollback() thay vi
            # chay den day — STB vua ve cung bi huy theo (full rollback).
            conn.commit()
            conn.close()
            conn = None

            # Xuat CSV CHI xay ra SAU KHI DB da commit thanh cong.
            steps_log.append({"step": "export", "status": "running"})
            export_path = os.path.join(os.path.dirname(self.db_path), "match_results.csv")
            try:
                export_match_results(result, export_path)
                steps_log.append({"step": "export", "status": "done", "detail": export_path})
            except Exception as e:
                steps_log.append({
                    "step": "export", "status": "error",
                    "detail": err("error_exporting_csv", detail=str(e)),
                })
                export_path = None

            return _ok({
                "steps": steps_log,
                "n_matched": n_matched,
                "n_total": len(result.assignment),
                "rounds_run": result.rounds_run,
                "export_path": export_path,
                "stb_redrawn": stb_redrawn,
            })
        except BaseException as e:
            # BaseException (khong chi Exception): KeyboardInterrupt/
            # SystemExit KHONG ke thua tu Exception, nhung mot "crash
            # giua chung" phai duoc rollback du no la loai nao — full
            # rollback (Option A) khong duoc phep chi ap dung cho mot
            # so loai loi. Sau khi don dep xong, neu day la tin hieu
            # dieu khien tien trinh (Ctrl+C, thoat) thi PHAI nem lai
            # (raise) thay vi nuot thanh {ok: False} nhu mot loi
            # nghiep vu binh thuong.
            if conn is not None:
                try:
                    conn.rollback()
                except Exception:
                    pass
                conn.close()
                steps_log.append({"step": "rollback", "status": "done", "detail": err("pipeline_rolled_back")})
            if not isinstance(e, Exception):
                raise
            error_entry = err("error_running_pipeline", detail=str(e))
            steps_log.append({"step": "unknown", "status": "error", "detail": error_entry})
            return _fail({"steps": steps_log, "errors": [error_entry, traceback.format_exc()]})

    # -----------------------------------------------------------------
    # NHẬP DỮ LIỆU TỪ MICROSOFT FORMS (CSV đã chuẩn hoá bởi
    # 06_ms_forms_transform.py) — trước đây KHÔNG có đường nào để CSV
    # này vào app.db, đây là mảnh còn thiếu duy nhất trong luồng dữ liệu.
    #
    # JS đọc file bằng FileReader.readAsText() ở phía trình duyệt rồi
    # gửi NGUYÊN VĂN nội dung CSV (chuỗi text) qua đây — không cần
    # main.py hỗ trợ thêm gì, không phụ thuộc hộp thoại chọn file gốc
    # của hệ điều hành (vốn không ổn định trong mọi bản pywebview).
    #
    # Hỗ trợ 2 định dạng cho MỖI loại CSV — tự nhận diện theo header,
    # không cần người dùng chọn định dạng:
    #   (a) "dài" (long) — đúng 1-1 với cấu trúc bảng DB, do
    #       06_ms_forms_transform.py xuất ra:
    #         nguyện vọng:      student_id,name,club_id,rank
    #         chọn club thi:    student_id,club_id
    #   (b) "rộng" (wide) — 1 dòng/học sinh, tiện đọc bằng mắt:
    #         nguyện vọng:      student_id,name,pref_1,pref_2,...,pref_10
    #         chọn club thi:    student_id,name,test_club_1,test_club_2,...
    # -----------------------------------------------------------------

    @staticmethod
    def _parse_csv_rows(csv_text: str):
        """Tự nhận diện dấu phân cách (, hoặc ;) và trả về (fieldnames, rows)."""
        # BOM (\ufeff) — Excel LUÔN thêm ký tự này vào đầu file khi lưu
        # "CSV UTF-8", và Microsoft Forms xuất ra cũng vậy. Nó vô hình
        # khi mở file bằng mắt, nhưng dính liền vào tên cột đầu tiên:
        # "student_id" đọc lên thành "\ufeffstudent_id". Không cắt bỏ thì
        # phần mềm báo THIẾU CỘT student_id trên một file hoàn toàn đúng —
        # lỗi cực khó hiểu với giáo viên. .strip() KHÔNG cắt được BOM vì
        # nó không phải khoảng trắng.
        csv_text = csv_text.lstrip("\ufeff")
        sample = csv_text[:4096]
        # CHI lay dau phan cach tu Sniffer, KHONG lay ca dialect.
        #
        # csv.Sniffer doan luon quy uoc trich dan, va no doan SAI: voi
        # 'HS1,"Tran ""Bo"" Van A, Jr.",clb_a' no tra ve doublequote=False,
        # tuc la bo quy uoc "" = mot dau nhay. Hau qua: o ten bi cat ngay
        # dau phay, phan duoi ('Jr."') roi sang cot ke ben va bi hieu la
        # ma club. Ca dong bi bo qua, chi kem mot canh bao "club khong ton
        # tai" khong lien quan gi toi nguyen nhan that.
        # Ten CLB tieng Viet rat hay co dau nhay: CLB "Vi Cong Dong".
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t").delimiter
        except csv.Error:
            delimiter = ","
        reader = csv.DictReader(io.StringIO(csv_text), delimiter=delimiter)
        fieldnames = [ (f or "").strip().lower() for f in (reader.fieldnames or []) ]
        rows = []
        for raw_row in reader:
            row = { (k or "").strip().lower(): (v.strip() if isinstance(v, str) else v)
                    for k, v in raw_row.items() if k is not None }
            rows.append(row)
        return fieldnames, rows

    # -----------------------------------------------------------------
    # TỰ NHẬN DIỆN LOẠI FILE
    #
    # Giao diện cũ có HAI ô nạp file và người dùng phải tự chọn đúng ô.
    # Kéo nhầm ô KHÔNG hề báo lỗi: file nguyện vọng dạng dài
    # (student_id, name, club_id, rank) nạp vào ô "chọn club thi" vẫn
    # khớp đủ cột, nên nó ghi thẳng vào club_test_selection và báo
    # "thành công". Nguyện vọng thật mất sạch mà không một cảnh báo —
    # lỗi làm sai kết quả phân bổ của cả trường mà không ai biết.
    #
    # Nên bỏ hẳn việc bắt người dùng chọn: đọc dòng tiêu đề là biết file
    # gì. Nhưng CHỈ kết luận khi CHẮC CHẮN — bộ cột
    # (student_id, club_id) không kèm rank vừa có thể là chọn club thi
    # dạng dài, vừa có thể là nguyện vọng dạng dài thiếu cột rank. Gặp
    # trường hợp đó thì HỎI LẠI, tuyệt đối không đoán: đoán sai là đúng
    # lại cái bug vừa chữa.
    # -----------------------------------------------------------------

    # -----------------------------------------------------------------
    # ĐỌC THẲNG FILE EXCEL (.xlsx)
    #
    # Microsoft Forms xuất kết quả ra .xlsx. Trước đây người vận hành
    # phải mở Excel -> File -> Save As -> chọn đúng "CSV UTF-8" -> rồi
    # mới nạp được. Bước thừa đó lại là bước dễ sai nhất: chọn nhầm
    # "CSV (Comma delimited)" thì tên tiếng Việt hỏng hết dấu, mà chọn
    # đúng thì Excel chèn BOM (bản cũ báo thiếu cột student_id vì thế).
    # Đọc thẳng .xlsx là bỏ được cả lớp lỗi đó.
    #
    # openpyxl là Python thuần, không cần .NET hay thư viện hệ thống —
    # khác hẳn pythonnet, nên không kéo theo rủi ro đóng gói đã gặp.
    # -----------------------------------------------------------------

    @staticmethod
    def _o_excel_thanh_chu(gia_tri) -> str:
        """Đưa một ô Excel về đúng chuỗi mà phần CSV đang chờ.

        Quan trọng nhất là số: Excel lưu chỉ tiêu 20 dưới dạng số thực,
        đọc thô ra thành "20.0" và int("20.0") ném ValueError — cả dòng
        CLB sẽ bị bỏ qua dù file hoàn toàn đúng.
        """
        if gia_tri is None:
            return ""
        if isinstance(gia_tri, bool):
            return "1" if gia_tri else ""
        if isinstance(gia_tri, float) and gia_tri.is_integer():
            return str(int(gia_tri))
        return str(gia_tri).strip()

    def xlsx_to_csv_text(self, file_base64: str, sheet_name: str = ""):
        """Đọc file .xlsx (gửi lên dạng base64) và trả về text CSV.

        Phần còn lại của luồng nhập giữ nguyên: detect_csv_kind và
        import_csv_auto vẫn làm việc trên text, không cần biết dữ liệu
        đến từ .csv hay .xlsx.
        """
        try:
            import base64 as _b64
            import csv as _csv
            import io as _io

            try:
                import openpyxl
            except ImportError:
                return _fail(err("xlsx_support_missing"))

            try:
                du_lieu = _b64.b64decode(file_base64 or "", validate=False)
                wb = openpyxl.load_workbook(
                    _io.BytesIO(du_lieu), read_only=True, data_only=True
                )
            except Exception as e:
                return _fail(err("xlsx_read_failed", detail=str(e)))

            ten_cac_sheet = list(wb.sheetnames)
            if not ten_cac_sheet:
                return _fail(err("xlsx_empty"))
            if sheet_name and sheet_name in ten_cac_sheet:
                ws = wb[sheet_name]
            else:
                # Sheet DAU TIEN, khong phai sheet dang active: file Forms
                # xuat ra luon de du lieu o sheet dau, con active co the la
                # sheet nguoi dung xem cuoi cung truoc khi luu.
                ws = wb[ten_cac_sheet[0]]

            dong = []
            for hang in ws.iter_rows(values_only=True):
                o = [self._o_excel_thanh_chu(v) for v in hang]
                # Excel hay de lai vai hang rong o cuoi bang.
                if any(x for x in o):
                    dong.append(o)
            wb.close()

            if not dong:
                return _fail(err("xlsx_empty"))

            buf = _io.StringIO()
            _csv.writer(buf, lineterminator="\n").writerows(dong)
            return _ok({
                "csv_text": buf.getvalue(),
                "sheet_name": ws.title,
                "sheet_names": ten_cac_sheet,
                "n_rows": len(dong),
            })
        except Exception as e:
            return _fail([err("xlsx_read_failed", detail=str(e)),
                          traceback.format_exc()])

    def detect_csv_kind(self, csv_text: str):
        """Đọc dòng tiêu đề để biết đây là file gì.

        Trả về {kind, format, confident, candidates, fieldnames}.
        confident=False nghĩa là giao diện PHẢI hỏi lại người dùng.
        """
        try:
            fieldnames, rows = self._parse_csv_rows(csv_text)
            if not rows:
                return _fail(err("csv_empty"))

            co = set(fieldnames)

            def ket_luan(kind, fmt):
                return _ok({"kind": kind, "format": fmt, "confident": True,
                            "candidates": [kind], "fieldnames": fieldnames})

            # capacity chi co nghia voi file danh sach CLB.
            if "capacity" in co and "club_id" in co:
                return ket_luan("clubs", "")
            if any(f.startswith("pref_") for f in fieldnames):
                return ket_luan("preferences", "wide")
            if any(f.startswith("test_club_") for f in fieldnames):
                return ket_luan("test_selection", "wide")
            # rank chi co nghia voi nguyen vong — chon club thi khong xep hang.
            if "rank" in co and "club_id" in co and "student_id" in co:
                return ket_luan("preferences", "long")

            if "student_id" in co and "club_id" in co:
                return _ok({
                    "kind": "", "format": "long", "confident": False,
                    "candidates": ["test_selection", "preferences"],
                    "fieldnames": fieldnames,
                })

            return _ok({"kind": "unknown", "format": "", "confident": False,
                        "candidates": [], "fieldnames": fieldnames})
        except Exception as e:
            return _fail(err("error_detecting_csv_kind", detail=str(e)))

    def import_csv_auto(self, csv_text: str, kind: str = "",
                        create_missing_students: bool = True):
        """Nạp một file CSV bất kỳ — tự nhận diện loại, không cần chọn ô.

        kind chỉ dùng khi tự nhận diện KHÔNG chắc: giao diện hỏi lại
        người dùng rồi truyền câu trả lời vào đây.
        """
        try:
            if not kind:
                nhan_dien = self.detect_csv_kind(csv_text)
                if not nhan_dien["ok"]:
                    return nhan_dien
                d = nhan_dien["data"]
                if not d["confident"]:
                    # KHONG doan. Tha khong nhap con hon nhap vao sai bang.
                    return _fail(err(
                        "csv_kind_ambiguous",
                        candidates=d["candidates"], fieldnames=d["fieldnames"],
                    ))
                kind = d["kind"]

            if kind == "clubs":
                res = self.import_clubs_csv(csv_text)
            elif kind == "preferences":
                res = self.import_preferences_csv(csv_text, create_missing_students)
            elif kind == "test_selection":
                res = self.import_test_selection_csv(csv_text, create_missing_students)
            else:
                return _fail(err("csv_kind_unknown"))

            if res["ok"]:
                res["data"]["kind"] = kind
            return res
        except Exception as e:
            return _fail([err("error_importing_csv_auto", detail=str(e)),
                          traceback.format_exc()])

    def import_clubs_csv(self, csv_text: str):
        """Nhập danh sách CLB bằng CSV.

        Đây là nút thắt ĐẦU TIÊN người dùng gặp: mẫu CSV học sinh bắt
        buộc club phải tồn tại trước (không thì cả học sinh bị bỏ qua),
        mà club lại chỉ tạo được bằng cách gõ form từng cái.

        Cột: club_id, name, capacity — bắt buộc.
             reserve_capacity, reserve_group — tuỳ chọn (trường không
             dùng dự trữ thì bỏ trống cả hai).
        """
        try:
            fieldnames, rows = self._parse_csv_rows(csv_text)
            if not rows:
                return _fail(err("csv_empty"))
            for bat_buoc in ("club_id", "name", "capacity"):
                if bat_buoc not in fieldnames:
                    return _fail(err("csv_missing_columns", fieldnames=fieldnames))

            conn = connect_db(self.db_path)
            cur = conn.cursor()
            da_co = {r[0] for r in cur.execute("SELECT club_id FROM clubs")}

            n_tao, n_sua, n_bo = 0, 0, 0
            canh_bao = []
            for i, row in enumerate(rows, start=2):   # dong 1 la tieu de
                club_id = (row.get("club_id") or "").strip()
                if not club_id:
                    canh_bao.append(err("csv_club_row_invalid", line=i,
                                        reason="club_id"))
                    n_bo += 1
                    continue
                try:
                    capacity = int(row.get("capacity") or 0)
                    reserve_capacity = int(row.get("reserve_capacity") or 0)
                except ValueError:
                    canh_bao.append(err("csv_club_row_invalid", line=i,
                                        reason="capacity"))
                    n_bo += 1
                    continue
                # Cung dung dieu kien nhu create_or_update_club — khong de
                # duong CSV lot qua thu duong UI da chan.
                if capacity <= 0 or reserve_capacity > capacity:
                    canh_bao.append(err("csv_club_row_invalid", line=i,
                                        reason="capacity"))
                    n_bo += 1
                    continue

                cur.execute(
                    """
                    INSERT INTO clubs (club_id, name, capacity, reserve_capacity, reserve_group)
                    VALUES (?, ?, ?, ?, ?)
                    ON CONFLICT(club_id) DO UPDATE SET
                        name=excluded.name,
                        capacity=excluded.capacity,
                        reserve_capacity=excluded.reserve_capacity,
                        reserve_group=excluded.reserve_group
                    """,
                    (club_id, (row.get("name") or "").strip(), capacity,
                     reserve_capacity,
                     self.chuan_hoa_nhom_du_tru(row.get("reserve_group")) or None),
                )
                if club_id in da_co:
                    n_sua += 1
                else:
                    n_tao += 1
                    da_co.add(club_id)

            conn.commit()
            conn.close()
            return _ok({
                "n_clubs_created": n_tao,
                "n_clubs_updated": n_sua,
                "n_rows_skipped": n_bo,
                "warnings": canh_bao,
            })
        except Exception as e:
            return _fail([err("error_importing_clubs_csv", detail=str(e)),
                          traceback.format_exc()])

    def preview_import_csv(self, csv_text: str, kind: str):
        """
        Xem trước trước khi nhập thật (không ghi DB) — cho UI hiện
        'phát hiện định dạng X, Y dòng, Z học sinh sẽ được tạo mới'
        trước khi người dùng bấm xác nhận nhập.
        kind: "preferences" hoặc "test_selection"
        """
        try:
            fieldnames, rows = self._parse_csv_rows(csv_text)
            if not rows:
                return _fail(err("csv_empty"))

            is_wide = any(f.startswith("pref_") or f.startswith("test_club_") for f in fieldnames)
            fmt = "wide" if is_wide else "long"

            conn = connect_db(self.db_path)
            existing_ids = {r[0] for r in conn.execute("SELECT student_id FROM students")}
            conn.close()

            row_student_ids = {r.get("student_id", "") for r in rows if r.get("student_id")}
            new_students = [sid for sid in row_student_ids if sid not in existing_ids]

            return _ok({
                "format": fmt,
                "kind": kind,
                "fieldnames": fieldnames,
                "n_rows": len(rows),
                "n_students_detected": len(row_student_ids),
                "n_new_students": len(new_students),
                "sample_row": rows[0] if rows else None,
            })
        except Exception as e:
            return _fail(err("error_reading_csv_preview", detail=str(e)))

    # -----------------------------------------------------------------
    # CHUAN HOA NHAN NHOM DU TRU
    #
    # reserve_group la chuoi TU DO, go o HAI noi (file danh sach CLB va
    # file hoc sinh) va phai khop nhau thi co che du tru moi chay. Truoc
    # day so khop bang chuoi chinh xac, nen CLB khai "chinh_sach" con
    # giao vien go "Chinh sach" la HAI NHOM KHAC NHAU: hoc sinh dien
    # chinh sach vao theo dien general, mat suat du tru. Pipeline chay
    # het va khong bao loi — chi co hai dong trong muc Canh bao du lieu,
    # rat de luot qua.
    #
    # Nay moi nhan deu di qua day TRUOC KHI ghi vao DB, nen hai ben go
    # kieu nao cung quy ve cung mot ma.
    #
    # Chuan hoa luc GHI chu khong phai luc SO SANH la co y:
    # rbda_priority_pipeline.py van so khop chuoi chinh xac nhu cu, khong
    # phai sua gi — cho nhay cam nhat cua du an khong bi dung toi.
    # -----------------------------------------------------------------

    @staticmethod
    def chuan_hoa_nhom_du_tru(raw) -> str:
        """"Chinh sach", "CHINH SACH", "Chinh-Sach" -> "chinh_sach"."""
        import re
        import unicodedata

        chu = (raw or "").strip()
        if not chu:
            return ""
        # NFD tach dau ra khoi nguyen am roi loc bo — nhung KHONG tach
        # duoc chu D gach ngang, phai thay tay truoc.
        chu = chu.replace("Đ", "D").replace("đ", "d")
        chu = "".join(
            c for c in unicodedata.normalize("NFD", chu)
            if not unicodedata.combining(c)
        )
        chu = re.sub(r"[^0-9a-zA-Z]+", "_", chu.lower())
        return chu.strip("_")

    # Cot reserve_group la TUY CHON trong ca hai file hoc sinh. Truoc day
    # hoc sinh tao bang CSV luon co nhom du tru rong, phai vao man hinh 04
    # gan tay tung em — ma nhom du tru CHINH LA co che uu tien cua RB-DA,
    # quen buoc do thi phan du tru vo hieu hoan toan va pipeline van chay
    # tron tru khong bao gi.
    #
    # Quy tac (ghi ca trong HUONG_DAN_CSV.md):
    #   o co gia tri -> GHI DE nhom hien co (nguoi nhap chu dong dua vao)
    #   o trong      -> GIU NGUYEN, khong xoa (file thieu cot khong duoc
    #                   lam mat du lieu da gan truoc do)
    @staticmethod
    def _doc_diem(chuoi: str):
        """Đọc một ô điểm. Trả về float, hoặc None nếu không phải số.

        Chấp nhận cả dấu phẩy thập phân: Excel bản tiếng Việt lưu 8,5
        chứ không phải 8.5. Chỉ đổi khi có ĐÚNG một dấu phẩy và không có
        dấu chấm nào — "1,234.5" là cách viết nghìn, không đụng vào.
        """
        chuoi = (chuoi or "").strip()
        if not chuoi:
            return None
        if chuoi.count(",") == 1 and "." not in chuoi:
            chuoi = chuoi.replace(",", ".")
        try:
            so = float(chuoi)
        except (TypeError, ValueError):
            return None
        return so if so == so and so not in (float("inf"), float("-inf")) else None

    @staticmethod
    def _soat_dong_trung(rows, is_wide: bool) -> list:
        """Cảnh báo khi một student_id xuất hiện nhiều lần trong file DẠNG RỘNG.

        Dạng rộng là mỗi học sinh một dòng, nên dòng thứ hai của cùng một
        mã sẽ âm thầm ghi đè dòng đầu — học sinh điền form hai lần, hoặc
        người nhập dán nhầm, là mất nguyện vọng của một dòng mà không ai
        biết. Dạng DÀI thì nhiều dòng mỗi học sinh là bình thường, không
        cảnh báo gì.
        """
        if not is_wide:
            return []
        dem: dict = {}
        for row in rows:
            sid = (row.get("student_id") or "").strip()
            if sid:
                dem[sid] = dem.get(sid, 0) + 1
        return [
            err("csv_duplicate_student_rows", student_id=sid, n=n)
            for sid, n in sorted(dem.items()) if n > 1
        ]

    @staticmethod
    def _soat_ma_trung_hoa_thuong(cur, ma_trong_file) -> list:
        """Cảnh báo khi mã học sinh mới chỉ khác hoa/thường với mã đã có.

        `hs001` và `HS001` tạo ra HAI học sinh riêng biệt: file tick chọn
        dùng kiểu này, file nguyện vọng dùng kiểu kia, thế là hai hồ sơ
        rời rạc mỗi cái thiếu một nửa, và pipeline xử lý như hai người.

        KHÔNG tự gộp — gộp nhầm hai em có thật là hỏng nặng hơn nhiều.
        Chỉ báo để người nhập tự quyết.
        """
        da_co = {
            r[0] for r in cur.execute("SELECT student_id FROM students").fetchall()
        }
        theo_thuong = {}
        for m in sorted(da_co):
            theo_thuong.setdefault(m.lower(), m)

        # Hai cach viet nam TRONG CUNG MOT FILE cung phai bat duoc: nguoi
        # nhap go tay 'hs201' o dong nay va 'HS201' o dong kia thi ca hai
        # deu la ma moi, khong ma nao co san trong CSDL de doi chieu.
        canh_bao = []
        for m in sorted(set(ma_trong_file)):
            if m in da_co:
                theo_thuong.setdefault(m.lower(), m)
                continue
            cu = theo_thuong.get(m.lower())
            if cu:
                canh_bao.append(
                    err("csv_student_id_case_conflict", student_id=m, da_co=cu)
                )
            else:
                theo_thuong[m.lower()] = m
        return canh_bao

    @staticmethod
    def _soat_ma_nghi_bi_cat(ma_trong_file) -> list:
        """Cảnh báo mã học sinh nghi bị Excel cắt mất số 0 đứng đầu.

        `0012345` để Excel tự nhận định dạng thì thành số `12345`. Phần
        mềm nhận đúng thứ Excel đã lưu nên KHÔNG cứu được — nhưng PHÁT
        HIỆN được: trong cùng một file, mã toàn chữ số mà ngắn hơn hẳn
        những mã còn lại gần như chắc chắn đã bị cắt.

        Chỉ cảnh báo, KHÔNG chặn nhập và KHÔNG tự thêm số 0 vào: phần
        mềm không có cách nào biết mã gốc dài bao nhiêu, đoán thêm là tự
        bịa dữ liệu.

        Ba điều kiện để tránh báo nhiễu:
          - chỉ xét mã TOÀN CHỮ SỐ (mã có chữ thì Excel không đụng tới)
          - cần ít nhất 3 mã như vậy mới đủ cơ sở nói cái nào bất thường
          - mã ngắn phải là THIỂU SỐ; phần lớn mã đều ngắn thì đó là quy
            ước của trường, không phải lỗi Excel
        """
        toan_so = [m for m in ma_trong_file if m.isdigit()]
        if len(toan_so) < 3:
            return []

        dem_do_dai: dict = {}
        for m in toan_so:
            dem_do_dai[len(m)] = dem_do_dai.get(len(m), 0) + 1
        # Do dai pho bien nhat; hoa thi lay do dai LON hon (an toan hon —
        # gia thiet ma day du moi la chuan).
        pho_bien = max(dem_do_dai.items(), key=lambda kv: (kv[1], kv[0]))[0]

        nghi = sorted({m for m in toan_so if len(m) < pho_bien})
        # Ma ngan chiem da so -> quy uoc cua truong, khong phai loi.
        if len(nghi) * 2 >= len(toan_so):
            return []

        return [
            err("csv_student_id_maybe_truncated",
                student_id=m, do_dai=len(m), do_dai_pho_bien=pho_bien)
            for m in nghi
        ]

    @staticmethod
    def _khop_club_id(cur):
        """Trả về hàm đưa club_id người dùng gõ về đúng mã gốc trong DB.

        Khớp CHÍNH XÁC trước; chỉ khi không thấy mới thử bỏ qua hoa/thường.
        Nếu trường thật sự tạo cả `clb_a` lẫn `CLB_A` thì mã khớp chính
        xác vẫn thắng — phần mềm không tự đoán hộ.

        Tha thứ hoa/thường KHÔNG phải tha thứ mọi thứ: mã sai hẳn vẫn bị
        bỏ qua kèm cảnh báo như cũ.
        """
        that = [r[0] for r in cur.execute("SELECT club_id FROM clubs").fetchall()]
        chinh_xac = set(that)
        theo_thuong = {}
        for m in that:
            theo_thuong.setdefault(m.lower(), m)

        def khop(cid: str):
            cid = (cid or "").strip()
            if cid in chinh_xac:
                return cid
            return theo_thuong.get(cid.lower())

        return khop

    @staticmethod
    def _soat_nhom_du_tru_la(cur, nhom_da_ghi: dict) -> list:
        """Cảnh báo NGAY LÚC NHẬP nếu nhãn nhóm không CLB nào nhận.

        Mục "Cảnh báo dữ liệu" cũng bắt được ca này, nhưng nằm ở panel
        khác và rất dễ lướt qua. Sai ở đây thì học sinh diện ưu tiên mất
        suất dự trữ mà pipeline vẫn chạy trơn, nên phải đập vào mắt ngay
        khi vừa nạp file.

        nhom_da_ghi: { nhãn đã chuẩn hoá: số học sinh mang nhãn đó }
        """
        import difflib

        if not nhom_da_ghi:
            return []
        nhom_clb = {
            r[0] for r in cur.execute(
                "SELECT DISTINCT reserve_group FROM clubs "
                "WHERE reserve_group IS NOT NULL AND TRIM(reserve_group) <> ''"
            ).fetchall()
        }
        canh_bao = []
        for nhan, n in sorted(nhom_da_ghi.items()):
            if nhan in nhom_clb:
                continue
            # Go nham mot hai ky tu la ca gap nhat — chi thang nhom dinh go.
            gan_giong = difflib.get_close_matches(nhan, sorted(nhom_clb), n=1, cutoff=0.6)
            canh_bao.append(err(
                "csv_reserve_group_unknown",
                reserve_group=nhan, n=n,
                goi_y=gan_giong[0] if gan_giong else "",
            ))
        return canh_bao

    @staticmethod
    def _ghi_nhom_du_tru(cur, student_id: str, reserve_group: str) -> str:
        """Trả về nhãn đã chuẩn hoá (chuỗi rỗng nếu không ghi gì) để bên
        gọi đếm được từng nhãn phục vụ _soat_nhom_du_tru_la."""
        nhom = PipelineAPI.chuan_hoa_nhom_du_tru(reserve_group)
        if not nhom:
            return ""
        cur.execute(
            "UPDATE students SET reserve_group = ? WHERE student_id = ?",
            (nhom, student_id),
        )
        return nhom

    def import_preferences_csv(self, csv_text: str, create_missing_students: bool = True):
        """
        Nhập CSV nguyện vọng (Bước 2 — xếp hạng) từ Microsoft Forms.
        Ghi đè TOÀN BỘ nguyện vọng cũ của TỪNG học sinh xuất hiện trong
        file (giống hành vi submit_preferences ở kiosk) — không đụng
        tới học sinh không có trong file.
        """
        try:
            fieldnames, rows = self._parse_csv_rows(csv_text)
            if not rows:
                return _fail(err("csv_empty"))

            is_wide = any(f.startswith("pref_") for f in fieldnames)

            # Gom thanh { student_id: (name, [club_id_theo_thu_tu]) }
            grouped: dict = {}
            if is_wide:
                pref_cols = sorted(
                    [f for f in fieldnames if f.startswith("pref_")],
                    key=lambda f: int(f.split("_")[1]) if f.split("_")[1].isdigit() else 999,
                )
                for row in rows:
                    sid = row.get("student_id")
                    if not sid:
                        continue
                    ordered = [row[c] for c in pref_cols if row.get(c)]
                    grouped[sid] = (row.get("name", ""), ordered,
                                    row.get("reserve_group", ""))
            else:
                if "student_id" not in fieldnames or "club_id" not in fieldnames:
                    return _fail(err("csv_missing_columns", fieldnames=fieldnames))
                has_rank = "rank" in fieldnames
                by_sid_rows: dict = {}
                for row in rows:
                    sid = row.get("student_id")
                    if not sid or not row.get("club_id"):
                        continue
                    by_sid_rows.setdefault(sid, []).append(row)
                for sid, sid_rows in by_sid_rows.items():
                    if has_rank:
                        sid_rows.sort(key=lambda r: int(r["rank"]) if r.get("rank", "").isdigit() else 999)
                    name = next((r.get("name") for r in sid_rows if r.get("name")), "")
                    nhom = next((r.get("reserve_group") for r in sid_rows
                                 if r.get("reserve_group")), "")
                    grouped[sid] = (name, [r["club_id"] for r in sid_rows], nhom)

            conn = connect_db(self.db_path)
            cur = conn.cursor()
            khop_club = self._khop_club_id(cur)
            existing_students = {r[0] for r in cur.execute("SELECT student_id FROM students")}

            n_created, n_updated, n_skipped = 0, 0, 0
            row_errors = self._soat_dong_trung(rows, is_wide)
            row_errors += self._soat_ma_trung_hoa_thuong(cur, grouped.keys())
            row_errors += self._soat_ma_nghi_bi_cat(grouped.keys())
            # Điểm chỉ thuộc file CHỌN CLB THI. Gặp cột điểm ở đây là
            # người nhập tưởng đã nạp điểm rồi — im lặng bỏ qua đúng là
            # loại lỗi im lặng dự án này đã phải sửa nhiều lần.
            if any(f == "score" or f.startswith("score_") for f in fieldnames):
                row_errors.append(err("csv_scores_ignored_here"))
            nhom_da_ghi: dict = {}

            for sid, (name, ordered_clubs, nhom_du_tru) in grouped.items():
                # loai bo trung lap giu thu tu xuat hien dau tien
                seen = set()
                deduped = []
                for cid in ordered_clubs:
                    if cid not in seen:
                        seen.add(cid)
                        deduped.append(cid)
                if len(deduped) != len(ordered_clubs):
                    row_errors.append(err("csv_pref_duplicate_deduped", student_id=sid))

                if len(deduped) > 10:
                    row_errors.append(err("csv_pref_too_many_skipped", student_id=sid, count=len(deduped)))
                    n_skipped += 1
                    continue

                # Doi ma nguoi dung go ve dung ma GOC trong DB (chap nhan
                # khac hoa/thuong) truoc khi kiem tra ton tai.
                da_khop = [(c, khop_club(c)) for c in deduped]
                invalid_clubs = [c for c, that in da_khop if that is None]
                if invalid_clubs:
                    row_errors.append(err("csv_unknown_clubs_skipped", student_id=sid, club_ids=invalid_clubs))
                    n_skipped += 1
                    continue
                deduped = [that for _, that in da_khop]

                if sid not in existing_students:
                    if not create_missing_students:
                        row_errors.append(err("csv_student_missing_skipped", student_id=sid))
                        n_skipped += 1
                        continue
                    cur.execute(
                        "INSERT INTO students (student_id, name, stb_number, reserve_group) "
                        "VALUES (?, ?, NULL, NULL)",
                        (sid, name or sid),
                    )
                    existing_students.add(sid)
                    n_created += 1
                elif name:
                    cur.execute(
                        "UPDATE students SET name = ? WHERE student_id = ? AND (name IS NULL OR name = '')",
                        (name, sid),
                    )

                nhan = self._ghi_nhom_du_tru(cur, sid, nhom_du_tru)
                if nhan:
                    nhom_da_ghi[nhan] = nhom_da_ghi.get(nhan, 0) + 1

                cur.execute("DELETE FROM preferences WHERE student_id = ?", (sid,))
                cur.executemany(
                    "INSERT INTO preferences (student_id, club_id, rank) VALUES (?, ?, ?)",
                    [(sid, cid, i + 1) for i, cid in enumerate(deduped)],
                )
                n_updated += 1

            row_errors.extend(self._soat_nhom_du_tru_la(cur, nhom_da_ghi))

            conn.commit()
            conn.close()

            return _ok({
                "n_students_created": n_created,
                "n_students_with_preferences_written": n_updated,
                "n_students_skipped": n_skipped,
                "warnings": row_errors,
            })
        except Exception as e:
            return _fail([err("error_importing_preferences_csv", detail=str(e)), traceback.format_exc()])

    def import_test_selection_csv(self, csv_text: str, create_missing_students: bool = True):
        """
        Nhập CSV chọn club muốn thi/xét (Bước 1 — tick-box) từ
        Microsoft Forms. Ghi đè toàn bộ lựa chọn thi cũ của từng học
        sinh xuất hiện trong file, giống hành vi submit_test_selection.
        """
        try:
            fieldnames, rows = self._parse_csv_rows(csv_text)
            if not rows:
                return _fail(err("csv_empty"))

            is_wide = any(f.startswith("test_club_") for f in fieldnames)

            grouped: dict = {}
            diem_tho: dict = {}     # sid -> {club_id thô: chuỗi điểm}
            loi_som: list = []      # lỗi phát hiện trước khi mở CSDL

            def _hau_to(ten: str, tien_to: str):
                duoi = ten[len(tien_to):]
                return duoi if duoi.isdigit() else None

            if is_wide:
                # Ghép test_club_N với score_N theo HẬU TỐ SỐ trong tên
                # cột, KHÔNG theo vị trí. Ô CLB bỏ trống bị lọc đi, nên
                # ghép theo vị trí sẽ lệch: em bỏ trống test_club_2 mà
                # điền test_club_3 thì điểm sẽ gán nhầm cho club khác.
                cot_clb = {}
                cot_diem = {}
                for f in fieldnames:
                    if f.startswith("test_club_"):
                        n = _hau_to(f, "test_club_")
                        if n:
                            cot_clb[n] = f
                    elif f.startswith("score_"):
                        n = _hau_to(f, "score_")
                        if n:
                            cot_diem[n] = f

                for row in rows:
                    sid = row.get("student_id")
                    if not sid:
                        continue
                    selected, diem_hs = [], {}
                    for n in sorted(cot_clb, key=int):
                        cid = (row.get(cot_clb[n]) or "").strip()
                        cot_d = cot_diem.get(n)
                        d = (row.get(cot_d) or "").strip() if cot_d else ""
                        if cid:
                            selected.append(cid)
                            if d:
                                diem_hs[cid] = d
                        elif d:
                            # Có điểm mà không có club -> gần như chắc là
                            # gõ lệch cột. Báo, đừng đoán club nào.
                            loi_som.append(err("csv_score_without_club",
                                               student_id=sid, cot="score_" + n))
                    grouped[sid] = (row.get("name", ""), selected,
                                    row.get("reserve_group", ""))
                    if diem_hs:
                        diem_tho[sid] = diem_hs
            else:
                if "student_id" not in fieldnames or "club_id" not in fieldnames:
                    return _fail(err("csv_missing_columns", fieldnames=fieldnames))
                for row in rows:
                    sid = row.get("student_id")
                    if not sid or not row.get("club_id"):
                        continue
                    name, clubs_list, nhom = grouped.get(sid, ("", [], ""))
                    grouped[sid] = (row.get("name") or name,
                                    clubs_list + [row["club_id"]],
                                    row.get("reserve_group") or nhom)
                    d = (row.get("score") or "").strip()
                    if d:
                        diem_tho.setdefault(sid, {})[row["club_id"]] = d

            conn = connect_db(self.db_path)
            cur = conn.cursor()
            khop_club = self._khop_club_id(cur)
            existing_students = {r[0] for r in cur.execute("SELECT student_id FROM students")}

            n_created, n_updated, n_skipped, n_diem = 0, 0, 0, 0
            row_errors = list(loi_som)
            row_errors += self._soat_dong_trung(rows, is_wide)
            row_errors += self._soat_ma_trung_hoa_thuong(cur, grouped.keys())
            row_errors += self._soat_ma_nghi_bi_cat(grouped.keys())
            nhom_da_ghi: dict = {}

            for sid, (name, club_ids, nhom_du_tru) in grouped.items():
                deduped = sorted(set(club_ids), key=club_ids.index)
                da_khop = [(c, khop_club(c)) for c in deduped]
                invalid_clubs = [c for c, that in da_khop if that is None]
                if invalid_clubs:
                    row_errors.append(err("csv_unknown_clubs_skipped", student_id=sid, club_ids=invalid_clubs))
                    n_skipped += 1
                    continue
                deduped = [that for _, that in da_khop]

                if sid not in existing_students:
                    if not create_missing_students:
                        row_errors.append(err("csv_student_missing_skipped", student_id=sid))
                        n_skipped += 1
                        continue
                    cur.execute(
                        "INSERT INTO students (student_id, name, stb_number, reserve_group) "
                        "VALUES (?, ?, NULL, NULL)",
                        (sid, name or sid),
                    )
                    existing_students.add(sid)
                    n_created += 1
                elif name:
                    cur.execute(
                        "UPDATE students SET name = ? WHERE student_id = ? AND (name IS NULL OR name = '')",
                        (name, sid),
                    )

                nhan = self._ghi_nhom_du_tru(cur, sid, nhom_du_tru)
                if nhan:
                    nhom_da_ghi[nhan] = nhom_da_ghi.get(nhan, 0) + 1

                cur.execute("DELETE FROM club_test_selection WHERE student_id = ?", (sid,))
                cur.executemany(
                    "INSERT INTO club_test_selection (student_id, club_id) VALUES (?, ?)",
                    [(sid, cid) for cid in deduped],
                )
                n_updated += 1

                # Điểm ghi SAU lựa chọn thi, và ghi bằng club_id ĐÃ KHỚP
                # (khop_club chấp nhận lệch hoa/thường), không phải chuỗi
                # thô trong file — ghi thô sẽ tạo một dòng điểm mồ côi mà
                # không màn hình nào đọc tới.
                diem_hs = diem_tho.get(sid)
                if diem_hs:
                    anh_xa = {tho: that for tho, that in da_khop}
                    for tho, chuoi in diem_hs.items():
                        that = anh_xa.get(tho)
                        if that is None:
                            row_errors.append(err("csv_score_for_unselected_club",
                                                  student_id=sid, club_id=tho))
                            continue
                        so = self._doc_diem(chuoi)
                        if so is None:
                            # Chỉ bỏ RIÊNG ô điểm này. Bỏ cả học sinh vì
                            # một ô gõ sai là mất nhiều hơn được.
                            row_errors.append(err("csv_score_not_a_number",
                                                  student_id=sid, club_id=that, score=chuoi))
                            continue
                        if so < 0:
                            row_errors.append(err("csv_score_negative",
                                                  student_id=sid, club_id=that, score=chuoi))
                            continue
                        cur.execute(
                            "INSERT INTO club_scores (student_id, club_id, score) VALUES (?, ?, ?) "
                            "ON CONFLICT(student_id, club_id) DO UPDATE SET score = excluded.score",
                            (sid, that, so),
                        )
                        n_diem += 1

            row_errors.extend(self._soat_nhom_du_tru_la(cur, nhom_da_ghi))

            conn.commit()
            conn.close()

            return _ok({
                "n_students_created": n_created,
                "n_students_with_selection_written": n_updated,
                "n_scores_written": n_diem,
                "n_students_skipped": n_skipped,
                "warnings": row_errors,
            })
        except Exception as e:
            return _fail([err("error_importing_test_selection_csv", detail=str(e)), traceback.format_exc()])

    # -----------------------------------------------------------------
    # CHẤM ĐIỂM MÙ (blind scoring) — trước đây club_scores chỉ được
    # điền qua seed_sample_data(), KHÔNG có màn hình cho giáo viên chấm
    # thật. Yêu cầu thiết kế cốt lõi "chấm mù" (grader không thấy STB,
    # không thấy thứ hạng nguyện vọng của học sinh) được ĐẢM BẢO Ở ĐÂY
    # bằng cách chỉ trả về student_id + tên — KHÔNG BAO GIỜ trả stb_number
    # hay preferences trong bất kỳ hàm nào của mục này.
    # -----------------------------------------------------------------

    def get_scoring_overview(self):
        """Tổng quan tiến độ chấm điểm theo từng club — cho màn hình chọn club để chấm."""
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            rows = cur.execute("""
                SELECT c.club_id, c.name,
                       COUNT(DISTINCT t.student_id) AS n_applicants,
                       COUNT(DISTINCT sc.student_id) AS n_scored
                FROM clubs c
                LEFT JOIN club_test_selection t ON t.club_id = c.club_id
                LEFT JOIN club_scores sc ON sc.club_id = c.club_id AND sc.student_id = t.student_id
                GROUP BY c.club_id
                ORDER BY c.club_id
            """).fetchall()
            conn.close()
            return _ok([dict(r) for r in rows])
        except Exception as e:
            return _fail(err("error_reading_scoring_overview", detail=str(e)))

    def get_club_applicants_for_scoring(self, club_id: str):
        """
        Danh sách học sinh cần chấm cho 1 club — CHỈ mã số, tên, và
        điểm đã chấm trước đó (nếu có, để sửa). KHÔNG kèm STB, KHÔNG
        kèm thứ hạng nguyện vọng — đúng yêu cầu chấm mù (blind scoring).
        """
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            club = cur.execute(
                "SELECT club_id, name FROM clubs WHERE club_id = ?", (club_id,)
            ).fetchone()
            if not club:
                conn.close()
                return _fail(err("club_not_found", club_id=club_id))
            rows = cur.execute("""
                SELECT s.student_id, s.name, sc.score
                FROM club_test_selection t
                JOIN students s ON s.student_id = t.student_id
                LEFT JOIN club_scores sc ON sc.student_id = t.student_id AND sc.club_id = t.club_id
                WHERE t.club_id = ?
                ORDER BY s.student_id
            """, (club_id,)).fetchall()
            conn.close()
            return _ok({
                "club_id": club["club_id"],
                "club_name": club["name"],
                "applicants": [dict(r) for r in rows],
            })
        except Exception as e:
            return _fail(err("error_reading_scoring_list", detail=str(e)))

    def submit_club_scores(self, club_id: str, scores: list):
        """
        Lưu điểm chấm cho 1 club. scores: [{"student_id": "...", "score": 8.5}, ...]
        Chỉ cho điểm học sinh THỰC SỰ có trong club_test_selection của
        club này (không thể chấm 'khống' cho học sinh không thi/xét).
        """
        try:
            if not isinstance(scores, list) or not scores:
                return _fail(err("scores_must_be_nonempty_list"))

            conn = connect_db(self.db_path)
            cur = conn.cursor()
            club_exists = cur.execute(
                "SELECT 1 FROM clubs WHERE club_id = ?", (club_id,)
            ).fetchone()
            if not club_exists:
                conn.close()
                return _fail(err("club_not_found", club_id=club_id))

            valid_applicants = {
                r[0] for r in cur.execute(
                    "SELECT student_id FROM club_test_selection WHERE club_id = ?",
                    (club_id,),
                ).fetchall()
            }

            n_saved, skipped = 0, []
            for entry in scores:
                sid = entry.get("student_id")
                score = entry.get("score")
                if sid not in valid_applicants:
                    skipped.append(err("score_not_applicant", student_id=sid))
                    continue
                if score is None or score == "":
                    # cho phep xoa diem (bo trong o) bang cach xoa ban ghi
                    cur.execute(
                        "DELETE FROM club_scores WHERE student_id = ? AND club_id = ?",
                        (sid, club_id),
                    )
                    continue
                # Dùng CHUNG bộ đọc số với đường nạp tệp, thay vì float()
                # thẳng. `_doc_diem` nhận cả dấu phẩy thập phân kiểu Việt
                # ("8,5") lẫn dấu chấm, và loại inf/nan. Trước đây hai cửa
                # hai luật: nạp tệp hiểu "8,5" là 8.5, còn màn hình chấm
                # điểm thì không — mà đó mới là cửa giáo viên gõ tay.
                so = self._doc_diem(score if isinstance(score, str) else str(score))
                if so is None:
                    skipped.append(err("score_not_a_number", student_id=sid, score=score))
                    continue
                score = so
                # Điểm âm: đường NẠP TỆP đã từ chối từ trước
                # (csv_score_negative), nhưng màn hình chấm điểm thì nhận.
                # Cùng một lỗi thừa dấu trừ, bắt được hay không lại tuỳ
                # giáo viên đi cửa nào — đo được: -9 vào lọt ở đây, và em
                # điểm cao nhất tụt xuống dưới tất cả, mất chỗ.
                if score < 0:
                    skipped.append(err("score_negative", student_id=sid, score=score))
                    continue
                cur.execute(
                    "INSERT INTO club_scores (student_id, club_id, score) VALUES (?, ?, ?) "
                    "ON CONFLICT(student_id, club_id) DO UPDATE SET score = excluded.score",
                    (sid, club_id, score),
                )
                n_saved += 1

            conn.commit()
            conn.close()
            return _ok({"club_id": club_id, "n_saved": n_saved, "warnings": skipped})
        except Exception as e:
            return _fail(err("error_saving_scores", detail=str(e)))

    # -----------------------------------------------------------------
    # TAB "KẾT QUẢ"
    # -----------------------------------------------------------------

    def get_match_results(self, search: str = ""):
        """Bảng kết quả, lọc theo student_id/name nếu search có giá trị."""
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            query = """
                SELECT m.student_id, s.name, m.club_id, c.name as club_name,
                       m.matched_tier, m.rank_in_student_pref
                FROM match_results m
                JOIN students s ON s.student_id = m.student_id
                LEFT JOIN clubs c ON c.club_id = m.club_id
            """
            params = ()
            if search:
                query += " WHERE m.student_id LIKE ? OR s.name LIKE ?"
                params = (f"%{search}%", f"%{search}%")
            query += " ORDER BY m.student_id"
            rows = cur.execute(query, params).fetchall()
            conn.close()
            return _ok([dict(r) for r in rows])
        except Exception as e:
            return _fail(err("error_reading_results", detail=str(e)))

    def get_club_fill_stats(self):
        """Tỉ lệ lấp đầy mỗi club — dùng vẽ thanh progress bar."""
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            # matched_reserve: SỐ EM THỰC SỰ VÀO BẰNG SUẤT DỰ TRỮ — khác
            # với reserve_capacity là chỉ tiêu dự trữ của CLB. Biểu đồ
            # trước đây vẽ theo chỉ tiêu, tức là vẽ một thuộc tính của
            # CLB chứ không phải điều đã xảy ra. Dữ liệu vốn có sẵn ở
            # match_results.matched_tier, chỉ là câu lệnh chưa lấy.
            rows = cur.execute("""
                SELECT c.club_id, c.name, c.capacity, c.reserve_capacity,
                       COUNT(m.student_id) as matched,
                       SUM(CASE WHEN m.matched_tier = 'reserve' THEN 1 ELSE 0 END)
                           AS matched_reserve
                FROM clubs c
                LEFT JOIN match_results m ON m.club_id = c.club_id
                GROUP BY c.club_id
                ORDER BY c.club_id
            """).fetchall()
            conn.close()
            return _ok([dict(r) for r in rows])
        except Exception as e:
            return _fail(err("error_reading_club_stats", detail=str(e)))

    # -----------------------------------------------------------------
    # XUẤT KẾT QUẢ
    #
    # File này là SẢN PHẨM CUỐI của cả quy trình — nhà trường dán bảng,
    # gửi phụ huynh, phát cho giáo viên phụ trách từng CLB. Bản cũ chỉ
    # xuất hai cột mã (student_id,club_id): nhìn vào không biết em nào
    # tên gì, đỗ CLB nào, đỗ nguyện vọng thứ mấy. Toàn bộ dữ liệu đó ĐÃ
    # có sẵn trong DB, chỉ là câu lệnh xuất không lấy.
    # -----------------------------------------------------------------

    # Ký tự KHÔNG được phép có trong tên file trên Windows, cộng thêm
    # dấu phân cách đường dẫn. club_id do trường tự đặt và
    # create_or_update_club KHÔNG giới hạn ký tự, nên một mã như
    # "../ngoai" hay "khoi 10/11" sẽ làm file rơi ra ngoài thư mục kết
    # quả (hoặc ghi đè file khác) nếu ghép thẳng vào tên file.
    _KY_TU_CAM_TRONG_TEN_FILE = '<>:"/\\|?*'

    @staticmethod
    def _ten_file_an_toan(raw: str, mac_dinh: str = "club") -> str:
        ten = "".join(
            "_" if (c in PipelineAPI._KY_TU_CAM_TRONG_TEN_FILE or ord(c) < 32) else c
            for c in (raw or "")
        ).strip(" .")
        # "..", "." và chuỗi rỗng đều không dùng làm tên file được
        return ten if ten and ten not in (".", "..") else mac_dinh

    def export_csv(self, output_path: str = ""):
        """
        Xuất kết quả phân bổ ra CSV cho nhà trường:

          1. MỘT FILE TỔNG — mọi học sinh, đủ tên và diện trúng tuyển.
          2. MỘT THƯ MỤC theo CLB — mỗi CLB một file để phát cho giáo
             viên phụ trách, kèm file `_chua_duoc_xep.csv` liệt kê các
             em chưa vào CLB nào (nhóm nhà trường cần xử lý tiếp).

        output_path rỗng hoặc là đường dẫn TƯƠNG ĐỐI thì file được đặt
        CẠNH app.db. Đường dẫn tương đối vốn rơi vào thư mục làm việc
        của tiến trình — trên máy Windows chạy .exe qua shortcut, thư
        mục đó có thể là bất kỳ đâu, người dùng không tìm ra file. Luôn
        trả về đường dẫn ĐẦY ĐỦ để giao diện hiện đúng chỗ file nằm.
        """
        try:
            import csv

            output_path = (output_path or "").strip() or "ket_qua_phan_bo.csv"
            if not os.path.isabs(output_path):
                output_path = os.path.join(
                    os.path.dirname(os.path.abspath(self.db_path)),
                    os.path.basename(output_path),
                )

            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            rows = cur.execute("""
                SELECT m.student_id, s.name AS ho_ten, m.club_id,
                       c.name AS ten_club, m.rank_in_student_pref, m.matched_tier,
                       s.reserve_group
                FROM match_results m
                LEFT JOIN students s ON s.student_id = m.student_id
                LEFT JOIN clubs   c ON c.club_id   = m.club_id
                ORDER BY m.student_id
            """).fetchall()
            conn.close()

            def an_toan_cho_excel(o):
                """Chan Excel hieu noi dung o thanh CONG THUC.

                Excel tinh moi o bat dau bang = + - @ nhu cong thuc. Mot
                hoc sinh ten "=1+1" se hien ra la 2, va giao vien khong
                the biet ten that la gi. Them dau nhay don o dau la cach
                chuan de Excel hieu "day la chu, dung tinh".
                """
                if isinstance(o, str) and o[:1] in ("=", "+", "-", "@"):
                    return "'" + o
                return o

            def dien(tier):
                # 'reserve'/'general' la ma noi bo — giao vien khong phai doan.
                return {"reserve": "Dự trữ", "general": "Thường"}.get(tier or "", "")

            COT_TONG = ["Mã học sinh", "Họ tên", "Mã CLB", "Tên CLB",
                        "Nguyện vọng thứ", "Diện trúng tuyển", "Nhóm dự trữ"]
            COT_CLB = ["Mã học sinh", "Họ tên", "Nguyện vọng thứ",
                       "Diện trúng tuyển", "Nhóm dự trữ"]

            def dong_tong(r):
                return [
                    r["student_id"], r["ho_ten"] or "", r["club_id"] or "",
                    r["ten_club"] or ("" if r["club_id"] else "(chưa được xếp)"),
                    r["rank_in_student_pref"] if r["rank_in_student_pref"] else "",
                    dien(r["matched_tier"]), r["reserve_group"] or "",
                ]

            # encoding="utf-8-sig" = UTF-8 CÓ BOM. Không có BOM thì Excel
            # đọc tên tiếng Việt thành "Nguyá»…n VÄƒn An". Đây đúng là lỗi
            # đã gặp ở chiều NHẬP, chỉ ngược chiều.
            def ghi(path, header, cac_dong):
                with open(path, "w", newline="", encoding="utf-8-sig") as f:
                    w = csv.writer(f)
                    w.writerow(header)
                    w.writerows([an_toan_cho_excel(o) for o in d] for d in cac_dong)

            ghi(output_path, COT_TONG, [dong_tong(r) for r in rows])

            goc, _ = os.path.splitext(output_path)
            per_club_dir = goc + "_theo_club"
            os.makedirs(per_club_dir, exist_ok=True)

            # Xoa tep .csv cua lan xuat TRUOC trong dung thu muc nay.
            # Khong xoa thi mot CLB da bi go khoi he thong van con nguyen
            # tep cua no, nam canh cac tep moi va trong y het nhu that.
            # Giao vien cam nham tep do di to chuc mot CLB khong con ton
            # tai — sai ma khong co dau hieu nao bao.
            # Chi dung .csv, chi trong thu muc do phan mem tu tao ra.
            for cu_ten in os.listdir(per_club_dir):
                if cu_ten.lower().endswith(".csv"):
                    try:
                        os.remove(os.path.join(per_club_dir, cu_ten))
                    except OSError:
                        pass

            theo_club: dict = {}
            for r in rows:
                theo_club.setdefault(r["club_id"] or None, []).append(r)

            n_file = 0
            ten_da_dung: dict = {}
            for club_id, ds in theo_club.items():
                if club_id is None:
                    ten = "_chua_duoc_xep"
                else:
                    ten = self._ten_file_an_toan(club_id)
                    # Hai club_id khac nhau co the ra cung ten sau khi lam
                    # sach ("a/b" va "a\b") -> khong duoc de ghi de nhau.
                    if ten in ten_da_dung:
                        ten_da_dung[ten] += 1
                        ten = f"{ten}_{ten_da_dung[ten]}"
                    else:
                        ten_da_dung[ten] = 1
                ghi(
                    os.path.join(per_club_dir, ten + ".csv"),
                    COT_CLB,
                    [[r["student_id"], r["ho_ten"] or "",
                      r["rank_in_student_pref"] if r["rank_in_student_pref"] else "",
                      dien(r["matched_tier"]), r["reserve_group"] or ""] for r in ds],
                )
                n_file += 1

            return _ok({
                "path": output_path,
                "n_rows": len(rows),
                "per_club_dir": per_club_dir,
                "n_club_files": n_file,
            })
        except Exception as e:
            return _fail(err("error_exporting_csv", detail=str(e)))

    # -----------------------------------------------------------------
    # TAB "QUẢN LÝ CLUB & DỰ TRỮ" (admin thao tác trực tiếp — trường tự quyết)
    # -----------------------------------------------------------------

    def list_clubs_admin(self):
        """Giống list_clubs nhưng KÈM reserve_group, cho màn hình quản lý.

        Trước đây hàm này chỉ `return self.list_clubs()`, mà list_clubs
        không chọn cột reserve_group — nên bảng quản lý club luôn hiện
        "—" ở cột nhóm dự trữ dù DB có đủ dữ liệu. Nhóm dự trữ là cơ chế
        cốt lõi của RB-DA, hiển thị sai ở đây dễ khiến người vận hành
        tưởng chưa gán gì và đi cấu hình lại nhầm.
        """
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            rows = cur.execute(
                "SELECT club_id, name, capacity, reserve_capacity, reserve_group "
                "FROM clubs ORDER BY club_id"
            ).fetchall()
            conn.close()
            return _ok([dict(r) for r in rows])
        except Exception as e:
            return _fail(err("error_reading_club_list", detail=str(e)))

    def create_or_update_club(
        self, club_id: str, name: str, capacity: int,
        reserve_capacity: int = 0, reserve_group: str = "",
    ):
        """
        Tạo mới hoặc cập nhật 1 club (UPSERT theo club_id).
        reserve_group: chuỗi tự do do trường tự đặt (vd 'chinh_sach',
        'khoi10', hoặc để trống '' nếu club không có dự trữ).
        """
        try:
            capacity = int(capacity)
            reserve_capacity = int(reserve_capacity)
            if capacity <= 0:
                return _fail(err("capacity_must_be_positive"))
            if reserve_capacity > capacity:
                return _fail(err("reserve_capacity_exceeds_capacity"))
            if not club_id.strip():
                return _fail(err("club_id_required"))

            reserve_group_value = self.chuan_hoa_nhom_du_tru(reserve_group) or None

            conn = connect_db(self.db_path)
            cur = conn.cursor()
            cur.execute(
                """
                INSERT INTO clubs (club_id, name, capacity, reserve_capacity, reserve_group)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(club_id) DO UPDATE SET
                    name=excluded.name,
                    capacity=excluded.capacity,
                    reserve_capacity=excluded.reserve_capacity,
                    reserve_group=excluded.reserve_group
                """,
                (club_id.strip(), name.strip(), capacity, reserve_capacity, reserve_group_value),
            )
            conn.commit()
            conn.close()
            return _ok({"club_id": club_id, "action": "upserted"})
        except Exception as e:
            return _fail(err("error_saving_club", detail=str(e)))

    def delete_club(self, club_id: str):
        """
        Xoá club — CHẶN nếu đã có preferences/match_results tham chiếu
        tới club này, để tránh mất dữ liệu học sinh đã nộp.
        """
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            n_prefs = cur.execute(
                "SELECT COUNT(*) FROM preferences WHERE club_id = ?", (club_id,)
            ).fetchone()[0]
            n_matches = cur.execute(
                "SELECT COUNT(*) FROM match_results WHERE club_id = ?", (club_id,)
            ).fetchone()[0]
            if n_prefs > 0 or n_matches > 0:
                conn.close()
                return _fail(err(
                    "cannot_delete_club_referenced",
                    club_id=club_id, n_prefs=n_prefs, n_matches=n_matches,
                ))
            cur.execute("DELETE FROM clubs WHERE club_id = ?", (club_id,))
            cur.execute("DELETE FROM club_test_selection WHERE club_id = ?", (club_id,))
            cur.execute("DELETE FROM club_scores WHERE club_id = ?", (club_id,))
            conn.commit()
            conn.close()
            return _ok({"club_id": club_id, "deleted": True})
        except Exception as e:
            return _fail(err("error_deleting_club", detail=str(e)))

    def list_reserve_groups_in_use(self):
        """Danh sách các reserve_group đang được dùng (để gợi ý trong form, tránh gõ sai chính tả)."""
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            rows = cur.execute(
                "SELECT DISTINCT reserve_group FROM clubs WHERE reserve_group IS NOT NULL "
                "UNION "
                "SELECT DISTINCT reserve_group FROM students WHERE reserve_group IS NOT NULL"
            ).fetchall()
            conn.close()
            return _ok(sorted(r[0] for r in rows if r[0]))
        except Exception as e:
            return _fail(err("error_reading_reserve_groups", detail=str(e)))

    def set_student_reserve_group(self, student_id: str, reserve_group: str):
        """Gán (hoặc gỡ, nếu reserve_group='') diện dự trữ cho 1 học sinh."""
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            exists = cur.execute(
                "SELECT 1 FROM students WHERE student_id = ?", (student_id,)
            ).fetchone()
            if not exists:
                conn.close()
                return _fail(err("student_not_found", student_id=student_id))
            value = self.chuan_hoa_nhom_du_tru(reserve_group) or None
            cur.execute(
                "UPDATE students SET reserve_group = ? WHERE student_id = ?",
                (value, student_id),
            )
            conn.commit()
            conn.close()
            return _ok({"student_id": student_id, "reserve_group": value})
        except Exception as e:
            return _fail(err("error_assigning_reserve_group", detail=str(e)))

    def bulk_set_reserve_group(self, student_ids: list, reserve_group: str):
        """Gán hàng loạt — dùng khi trường có sẵn danh sách (vd cả 1 khối lớp)."""
        try:
            if not isinstance(student_ids, list) or not student_ids:
                return _fail(err("student_ids_must_be_nonempty_list"))
            value = self.chuan_hoa_nhom_du_tru(reserve_group) or None
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            existing_ids = {
                r[0] for r in cur.execute("SELECT student_id FROM students").fetchall()
            }
            missing = [sid for sid in student_ids if sid not in existing_ids]
            valid_ids = [sid for sid in student_ids if sid in existing_ids]
            cur.executemany(
                "UPDATE students SET reserve_group = ? WHERE student_id = ?",
                [(value, sid) for sid in valid_ids],
            )
            conn.commit()
            conn.close()
            return _ok({"n_updated": len(valid_ids), "not_found": missing})
        except Exception as e:
            return _fail(err("error_bulk_assigning", detail=str(e)))

    def list_students_admin(self, search: str = "", page: int = 1, page_size: int = 100):
        """
        Danh sách học sinh kèm reserve_group hiện tại, cho tab quản lý.
        CÓ PHÂN TRANG (trước đây LIMIT 100 cứng khiến trường >100 học
        sinh bị ẩn âm thầm không báo). Trả kèm total/total_pages để UI
        vẽ nút điều hướng trang.
        """
        try:
            page = max(1, int(page))
            page_size = max(1, min(int(page_size), 500))
            offset = (page - 1) * page_size

            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()

            where_clause = ""
            params: tuple = ()
            if search:
                where_clause = " WHERE student_id LIKE ? OR name LIKE ?"
                params = (f"%{search}%", f"%{search}%")

            total = cur.execute(
                f"SELECT COUNT(*) FROM students{where_clause}", params
            ).fetchone()[0]

            rows = cur.execute(
                f"SELECT student_id, name, reserve_group FROM students{where_clause} "
                "ORDER BY student_id LIMIT ? OFFSET ?",
                params + (page_size, offset),
            ).fetchall()
            conn.close()

            total_pages = max(1, (total + page_size - 1) // page_size)
            return _ok({
                "rows": [dict(r) for r in rows],
                "page": page,
                "page_size": page_size,
                "total": total,
                "total_pages": total_pages,
            })
        except Exception as e:
            return _fail(err("error_reading_student_list", detail=str(e)))

    # -----------------------------------------------------------------
    # TAB "NHẬP DỰ PHÒNG" (kiosk fallback entry)
    # -----------------------------------------------------------------

    def list_clubs(self):
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            rows = cur.execute(
                "SELECT club_id, name, capacity, reserve_capacity FROM clubs ORDER BY club_id"
            ).fetchall()
            conn.close()
            return _ok([dict(r) for r in rows])
        except Exception as e:
            return _fail(err("error_reading_club_list", detail=str(e)))

    def search_students(self, query: str):
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            rows = cur.execute(
                "SELECT student_id, name FROM students "
                "WHERE student_id LIKE ? OR name LIKE ? LIMIT 20",
                (f"%{query}%", f"%{query}%"),
            ).fetchall()
            conn.close()
            return _ok([dict(r) for r in rows])
        except Exception as e:
            return _fail(err("error_searching_students", detail=str(e)))

    def get_student_entry_state(self, student_id: str):
        """
        Trạng thái nhập liệu hiện tại của 1 học sinh — dùng để tab
        'Nhập dự phòng' hiển thị lại nếu học sinh quay lại kiosk.
        """
        try:
            conn = connect_db(self.db_path)
            conn.row_factory = sqlite3.Row
            cur = conn.cursor()
            student = cur.execute(
                "SELECT student_id, name FROM students WHERE student_id = ?",
                (student_id,),
            ).fetchone()
            if not student:
                conn.close()
                return _fail(err("student_not_found", student_id=student_id))

            tested = [
                r["club_id"]
                for r in cur.execute(
                    "SELECT club_id FROM club_test_selection WHERE student_id = ?",
                    (student_id,),
                ).fetchall()
            ]
            prefs = [
                r["club_id"]
                for r in cur.execute(
                    "SELECT club_id FROM preferences WHERE student_id = ? ORDER BY rank",
                    (student_id,),
                ).fetchall()
            ]
            conn.close()
            return _ok({
                "student_id": student["student_id"],
                "name": student["name"],
                "tested_clubs": tested,
                "ranked_clubs": prefs,
            })
        except Exception as e:
            return _fail(err("error_reading_student_state", detail=str(e)))

    def submit_test_selection(self, student_id: str, club_ids: list):
        """
        BƯỚC 1 (độc lập với xếp hạng) — tick-box club muốn thi/xét.
        Ghi đè toàn bộ lựa chọn cũ của học sinh này.
        """
        try:
            if not isinstance(club_ids, list):
                return _fail(err("club_ids_must_be_list"))
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            exists = cur.execute(
                "SELECT 1 FROM students WHERE student_id = ?", (student_id,)
            ).fetchone()
            if not exists:
                conn.close()
                return _fail(err("student_not_found", student_id=student_id))

            invalid = [
                cid for cid in club_ids
                if not cur.execute(
                    "SELECT 1 FROM clubs WHERE club_id = ?", (cid,)
                ).fetchone()
            ]
            if invalid:
                conn.close()
                return _fail(err("unknown_clubs", club_ids=invalid))

            cur.execute(
                "DELETE FROM club_test_selection WHERE student_id = ?", (student_id,)
            )
            cur.executemany(
                "INSERT INTO club_test_selection (student_id, club_id) VALUES (?, ?)",
                [(student_id, cid) for cid in club_ids],
            )
            conn.commit()
            conn.close()
            return _ok({"student_id": student_id, "n_selected": len(club_ids)})
        except Exception as e:
            return _fail(err("error_saving_test_selection", detail=str(e)))

    def submit_preferences(self, student_id: str, ordered_club_ids: list):
        """
        BƯỚC 2 (độc lập với tick-box thi) — xếp hạng nguyện vọng,
        tối đa 10 club (giới hạn Microsoft Forms Ranking, giữ đồng
        bộ với luồng nhập chính).
        """
        try:
            if not isinstance(ordered_club_ids, list):
                return _fail(err("ordered_club_ids_must_be_list"))
            if len(ordered_club_ids) == 0:
                return _fail(err("must_rank_at_least_one"))
            if len(ordered_club_ids) > 10:
                return _fail(err("max_10_preferences"))
            if len(ordered_club_ids) != len(set(ordered_club_ids)):
                return _fail(err("duplicate_preference_in_list"))

            conn = connect_db(self.db_path)
            cur = conn.cursor()
            exists = cur.execute(
                "SELECT 1 FROM students WHERE student_id = ?", (student_id,)
            ).fetchone()
            if not exists:
                conn.close()
                return _fail(err("student_not_found", student_id=student_id))

            invalid = [
                cid for cid in ordered_club_ids
                if not cur.execute(
                    "SELECT 1 FROM clubs WHERE club_id = ?", (cid,)
                ).fetchone()
            ]
            if invalid:
                conn.close()
                return _fail(err("unknown_clubs", club_ids=invalid))

            cur.execute("DELETE FROM preferences WHERE student_id = ?", (student_id,))
            cur.executemany(
                "INSERT INTO preferences (student_id, club_id, rank) VALUES (?, ?, ?)",
                [(student_id, cid, i + 1) for i, cid in enumerate(ordered_club_ids)],
            )
            conn.commit()
            conn.close()
            return _ok({"student_id": student_id, "n_ranked": len(ordered_club_ids)})
        except Exception as e:
            return _fail(err("error_saving_preferences", detail=str(e)))

    def create_student_if_missing(self, student_id: str, name: str):
        """Kiosk fallback: nếu học sinh chưa có trong students, tạo mới (chưa có STB)."""
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            exists = cur.execute(
                "SELECT 1 FROM students WHERE student_id = ?", (student_id,)
            ).fetchone()
            if not exists:
                cur.execute(
                    "INSERT INTO students (student_id, name, stb_number, reserve_group) "
                    "VALUES (?, ?, NULL, NULL)",
                    (student_id, name),
                )
                conn.commit()
            conn.close()
            return _ok({"student_id": student_id, "created": not exists})
        except Exception as e:
            return _fail(err("error_creating_student", detail=str(e)))

    def reset_student_entry(self, student_id: str):
        """
        Nút 'Sửa lại từ đầu' — xoá lựa chọn thi (Bước 1) và nguyện vọng
        (Bước 2) hiện tại của học sinh để nhập lại từ đầu tại kiosk.
        KHÔNG xoá bản ghi học sinh (giữ nguyên student_id/tên/STB/reserve_group).
        """
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            exists = cur.execute(
                "SELECT 1 FROM students WHERE student_id = ?", (student_id,)
            ).fetchone()
            if not exists:
                conn.close()
                return _fail(err("student_not_found", student_id=student_id))
            cur.execute("DELETE FROM club_test_selection WHERE student_id = ?", (student_id,))
            cur.execute("DELETE FROM preferences WHERE student_id = ?", (student_id,))
            conn.commit()
            conn.close()
            return _ok({"student_id": student_id, "reset": True})
        except Exception as e:
            return _fail(err("error_resetting_student_entry", detail=str(e)))

    def delete_student(self, student_id: str):
        """
        Nút 'Xoá học sinh' — xoá hẳn học sinh khỏi hệ thống (dùng khi
        tạo nhầm mã / trùng học sinh tại kiosk). CHẶN nếu học sinh đã
        có mặt trong match_results (đã qua lần chạy pipeline gần nhất)
        để tránh làm lệch thống kê lấp đầy club và mất dấu kiểm toán —
        phải chạy lại pipeline (hoặc xử lý ở tab Quản lý) trước.
        """
        try:
            conn = connect_db(self.db_path)
            cur = conn.cursor()
            exists = cur.execute(
                "SELECT 1 FROM students WHERE student_id = ?", (student_id,)
            ).fetchone()
            if not exists:
                conn.close()
                return _fail(err("student_not_found", student_id=student_id))
            n_matches = cur.execute(
                "SELECT COUNT(*) FROM match_results WHERE student_id = ?", (student_id,)
            ).fetchone()[0]
            if n_matches > 0:
                conn.close()
                return _fail(err("cannot_delete_student_matched", student_id=student_id))
            cur.execute("DELETE FROM club_test_selection WHERE student_id = ?", (student_id,))
            cur.execute("DELETE FROM preferences WHERE student_id = ?", (student_id,))
            cur.execute("DELETE FROM club_scores WHERE student_id = ?", (student_id,))
            cur.execute("DELETE FROM students WHERE student_id = ?", (student_id,))
            conn.commit()
            conn.close()
            return _ok({"student_id": student_id, "deleted": True})
        except Exception as e:
            return _fail(err("error_deleting_student", detail=str(e)))

    # -----------------------------------------------------------------
    # XOÁ DỮ LIỆU ĐỂ LÀM LẠI TỪ ĐẦU
    #
    # Nạp file CỘNG THÊM học sinh, không bao giờ xoá — đúng như phải thế,
    # vì trường nạp khối 10 rồi nạp khối 11 thì không được mất khối 10.
    # Nhưng trước đây KHÔNG có đường nào đưa dữ liệu về trống: cách duy
    # nhất là đóng app rồi đổi tên app.db ngoài File Explorer, việc mà
    # không ai tìm ra. Hậu quả không phải chuyện hình thức — học sinh của
    # lần chạy thử trước vẫn chiếm suất và làm lệch kết quả, im lặng.
    # -----------------------------------------------------------------

    _RESET_CONFIRM = "XOA"

    # Xoá theo đúng thứ tự phụ thuộc: bảng con trước, students sau cùng.
    _RESET_BANG = {
        "hoc_sinh": ["match_results", "club_scores", "preferences",
                     "club_test_selection", "students"],
        "tat_ca":   ["match_results", "club_scores", "preferences",
                     "club_test_selection", "students", "clubs"],
    }

    def reset_data(self, pham_vi: str = "hoc_sinh", xac_nhan: str = ""):
        """Xoá dữ liệu để chạy thử lại từ đầu. LUÔN sao lưu trước.

        pham_vi:
            "hoc_sinh" — xoá học sinh và mọi thứ gắn với học sinh, GIỮ
                         nguyên danh sách CLB (trường hợp thường gặp:
                         giữ CLB, thay khối học sinh khác vào).
            "tat_ca"   — xoá cả CLB.

        xac_nhan phải đúng chuỗi "XOA". Một hàm xoá sạch CSDL mà gọi
        nhầm được từ JS là chuyện không chấp nhận được.

        KHÔNG đụng vào run_history: lược đồ ghi rõ bảng đó "không bao giờ
        xoá/ghi đè" vì đó là nhật ký kiểm toán. Xoá dữ liệu không được
        phép xoá dấu vết đã từng chạy những gì.
        """
        try:
            # Kiểm tra TRƯỚC, sao lưu SAU. Đảo thứ tự thì một lần gọi
            # nhầm vẫn đẻ ra một tệp sao lưu rác mỗi lần.
            if xac_nhan != self._RESET_CONFIRM:
                return _fail(err("reset_confirmation_mismatch",
                                 can_go=self._RESET_CONFIRM))
            bang = self._RESET_BANG.get(pham_vi)
            if bang is None:
                return _fail(err("reset_scope_unknown", pham_vi=pham_vi,
                                 hop_le=", ".join(sorted(self._RESET_BANG))))

            backup_path = self._backup_db()

            conn = connect_db(self.db_path)
            cur = conn.cursor()
            # Đếm TRƯỚC khi xoá — báo sau khi xoá thì con số nào cũng 0.
            da_xoa = {
                ten: cur.execute("SELECT COUNT(*) FROM %s" % ten).fetchone()[0]
                for ten in bang
            }
            for ten in bang:
                cur.execute("DELETE FROM %s" % ten)

            # run_meta mô tả lần chạy gần nhất, mà lần chạy đó nay không
            # còn dữ liệu nào phía sau.
            cur.execute("DELETE FROM run_meta")

            # Mở khoá STB. Bỏ bước này thì lần chạy sau đi vào nhánh "đã
            # khoá" và ghi nhật ký là "tái sử dụng STB" cho một bộ số bốc
            # thăm không còn tồn tại — sai cho phần kiểm toán.
            cur.execute(
                "UPDATE stb_lock SET is_locked = 0, unlocked_at = ? WHERE id = 1",
                (_now(),),
            )
            # Đếm CLB CÒN LẠI ở đây chứ không để giao diện tự suy ra từ
            # phạm vi — suy ra là đoán, và đoán sai thì toast báo một con
            # số không ai kiểm chứng được.
            n_clb_con_lai = cur.execute("SELECT COUNT(*) FROM clubs").fetchone()[0]
            conn.commit()
            conn.close()

            return _ok({
                "pham_vi": pham_vi,
                "backup_path": backup_path,
                "backup_name": os.path.basename(backup_path),
                "da_xoa": da_xoa,
                "n_students": da_xoa.get("students", 0),
                "n_clubs": da_xoa.get("clubs", 0),
                "n_clubs_con_lai": n_clb_con_lai,
                "giu_lai_run_history": True,
            })
        except Exception as e:
            return _fail([err("error_resetting_data", detail=str(e)),
                          traceback.format_exc()])
